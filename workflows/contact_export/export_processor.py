#!/usr/bin/env python3
"""
Export Processor

Production-ready Contact Export workflow processor that uses unified core logic
and provides Contact Export-specific output formatting.
"""

try:
    import pandas as pd
except ImportError as e:
    print("ERROR: pandas could not be imported. Please ensure it's installed:")
    print("   pip install pandas>=1.5.0")
    print(f"   Import error: {e}")
    raise

import sys
import os
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError as e:
    print("ERROR: openpyxl could not be imported. Please ensure it's installed:")
    print("   pip install openpyxl>=3.0.0")
    print(f"   Import error: {e}")
    raise

# Add project root to path for imports
project_root = Path(__file__).parent.parent.parent
sys.path.insert(0, str(project_root))

from core.data_processor import DataProcessor
from core.llm_engine import LLMEngine
from core.name_parser import NameParser

class ExportProcessor:
    """
    Contact Export workflow processor
    
    Uses unified core logic from DataProcessor and provides Contact Export-specific
    output formatting and error tracking.
    """
    
    def __init__(self, config: Dict[str, Any]):
        self.config = config
        # Use the unified data processor for core logic
        self.data_processor = DataProcessor(config)
        # Keep direct access to LLM engine for workflow-specific needs
        # Transform config structure to match LLM engine expectations
        llm_config = {
            'openai': {
                'api_key': self._get_config_value('openai_api_key') or self._get_config_value('llm.openai_api_key'),
                'model': self._get_config_value('openai_model', 'gpt-4o-mini') or self._get_config_value('llm.openai_model', 'gpt-4o-mini'),
                'max_tokens': self._get_config_value('max_tokens', 150) or self._get_config_value('llm.max_tokens', 150),
                'temperature': self._get_config_value('temperature', 0.1) or self._get_config_value('llm.temperature', 0.1)
            },
            'ollama': {
                'base_url': self._get_config_value('ollama_base_url', 'http://localhost:11434') or self._get_config_value('llm.ollama_base_url', 'http://localhost:11434'),
                'model': self._get_config_value('ollama_model', 'qwen2.5:7b-instruct-q4_K_M') or self._get_config_value('llm.ollama_model', 'qwen2.5:7b-instruct-q4_K_M')
            }
        }
        self.llm_engine = LLMEngine(llm_config)
        self.name_parser = NameParser(self.llm_engine)
        
        # Store original source data for removed row tracking
        self._source_data = None
        self.original_columns = []
        
        # Performance optimization: lazy evaluation for removed rows
        self._removed_rows_cache = None
        self._removal_reasons = {}
        self._successfully_processed_emails = set()
    
    def _get_config_value(self, key_path, default=None):
        """Get config value from nested path, handling both dict and object access"""
        if isinstance(self.config, dict):
            # Handle dictionary access
            if '.' in key_path:
                parts = key_path.split('.')
                current = self.config
                for part in parts:
                    if isinstance(current, dict) and part in current:
                        current = current[part]
                    else:
                        return default
                return current
            else:
                return self.config.get(key_path, default)
        else:
            # Handle object access
            if '.' in key_path:
                parts = key_path.split('.')
                current = self.config
                for part in parts:
                    if hasattr(current, part):
                        current = getattr(current, part)
                    else:
                        return default
                return current
            else:
                return getattr(self.config, key_path, default)
    
    def process_contact_export(self, data: pd.DataFrame, original_source_data: pd.DataFrame = None, date_range: Optional[Dict[str, str]] = None) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Process contact export data using unified core logic with "Removed as Remaining" approach
        
        Args:
            data: Raw contact export data with proper column mapping
            original_source_data: Original source data before mapping (for removed row tracking)
            date_range: Optional date range filter
            
        Returns:
            Tuple of (processed_data, removed_rows)
        """
        print("Processing Contact Export using unified core logic with 'Removed as Remaining' approach...")
        
        # Store original source data for removed row tracking
        if original_source_data is not None:
            self._source_data = original_source_data.copy()
            self.original_columns = list(original_source_data.columns)
        else:
            self._source_data = data.copy()
            self.original_columns = list(data.columns)
        
        # Initialize removed rows with ALL source data (lazy evaluation for performance)
        self._removed_rows_cache = None  # Will be computed when needed
        self._removal_reasons = {}  # Track specific removal reasons
        self._successfully_processed_emails = set()  # Track successfully processed emails
        
        # Use unified core logic for data processing
        processed_data = self.data_processor.process_contact_export_data(data)
        
        # Apply Contact Export-specific post-processing
        processed_data, export_removed_count = self._apply_contact_export_specific_logic(processed_data)
        
        # Get final removed rows using performance optimization
        removed_rows = self._get_final_removed_rows()
        
        print(f"Contact Export processing complete using unified logic")
        print(f"   Records processed: {len(processed_data)}")
        print(f"   Records removed: {len(removed_rows)}")
        print(f"   Export-specific removals: {export_removed_count}")
        
        return processed_data, removed_rows
    
    def _apply_contact_export_specific_logic(self, df_processed: pd.DataFrame) -> Tuple[pd.DataFrame, int]:
        """
        Apply Contact Export-specific business logic and filtering.
        
        Args:
            df_processed: Processed data from core engine
            
        Returns:
            Tuple of (filtered_data, removal_count)
        """
        initial_count = len(df_processed)
        
        # Filter out "Accounts Receivable" salespeople
        ar_mask = df_processed['Salesperson'].str.contains('Accounts Receivable', case=False, na=False)
        ar_removed = df_processed[ar_mask]
        df_processed = df_processed[~ar_mask]
        
        # Store removal reasons for AR records
        for idx in ar_removed.index:
            self._removal_reasons[idx] = 'accounts_receivable_salesperson'
        
        # Filter out records with no valid person names
        name_mask = (
            (df_processed['first_name'].notna()) & 
            (df_processed['first_name'] != '') &
            (df_processed['last_name'].notna()) & 
            (df_processed['last_name'] != '')
        )
        no_name_removed = df_processed[~name_mask]
        df_processed = df_processed[name_mask]
        
        # Store removal reasons for no-name records
        for idx in no_name_removed.index:
            self._removal_reasons[idx] = 'missing_valid_person_names'
        
        # Remove duplicate emails (keep first occurrence) - use transformed column name
        email_column = 'email' if 'email' in df_processed.columns else 'Email'
        df_processed = df_processed.drop_duplicates(subset=[email_column], keep='first')
        
        # Track successfully processed emails for removed rows calculation
        self._successfully_processed_emails = set(df_processed[email_column].dropna())
        
        final_count = len(df_processed)
        total_removed = initial_count - final_count
        
        print(f"   Contact Export filtering: {initial_count} -> {final_count} records")
        print(f"   AR salespeople removed: {len(ar_removed)}")
        print(f"   No-name records removed: {len(no_name_removed)}")
        print(f"   Duplicate emails removed: {initial_count - len(df_processed.drop_duplicates(subset=[email_column], keep='first'))}")
        
        return df_processed, total_removed
    
    def _get_final_removed_rows(self) -> pd.DataFrame:
        """
        Get final removed rows using performance optimization.
        
        Returns:
            DataFrame of removed rows with removal reasons
        """
        if self._removed_rows_cache is None:
            # Use vectorized operation for O(n) performance
            # Use correct column name (transformed data has 'email')
            email_column = 'email' if 'email' in self._source_data.columns else 'Email'
            removed_mask = ~self._source_data[email_column].isin(self._successfully_processed_emails)
            self._removed_rows_cache = self._source_data[removed_mask].copy()
            
            # Add removal reasons column
            self._removed_rows_cache['Removal Reason'] = 'unknown'
            self._add_removal_reasons()
        
        return self._removed_rows_cache
    
    def _add_removal_reasons(self):
        """Add specific removal reasons to removed rows."""
        for idx, row in self._removed_rows_cache.iterrows():
            # Check if we have a tracked reason for this row
            if idx in self._removal_reasons:
                self._removed_rows_cache.at[idx, 'Removal Reason'] = self._removal_reasons[idx]
            else:
                # Analyze the row to determine removal reason
                reason = self._determine_specific_removal_reason(row)
                self._removed_rows_cache.at[idx, 'Removal Reason'] = reason
    
    def _determine_specific_removal_reason(self, row: pd.Series) -> str:
        """
        Determine specific removal reason for a row.
        
        Args:
            row: Row data to analyze
            
        Returns:
            Specific removal reason string
        """
        email = str(row.get('email', '')).strip()
        contact_name = str(row.get('contact_name', '')).strip()
        customer_name = str(row.get('customer_name', '')).strip()
        phone = str(row.get('phone', '')).strip()
        
        # Email-related reasons
        if not self._is_valid_email(email):
            if ';' in email:
                return 'multiple_emails_separated_by_semicolon'
            elif email.startswith('info@') or email.startswith('admin@') or email.startswith('support@'):
                return 'internal_system_email_address'
            elif '@' not in email or '.' not in email:
                return 'email_format_invalid'
            else:
                return 'email_validation_failed'
        
        # Name-related reasons
        if self._looks_like_business_name(contact_name) or self._looks_like_business_name(customer_name):
            return 'business_name_not_person'
        elif not contact_name and not customer_name:
            return 'missing_contact_and_customer_names'
        elif not contact_name:
            return 'missing_contact_name'
        elif not customer_name:
            return 'missing_customer_name'
        
        # Phone-related reasons
        if phone and not any(char.isdigit() for char in phone):
            return 'phone_number_invalid'
        
        # Default reason
        return 'processing_failed_unknown_reason'
    
    def _is_valid_email(self, email: str) -> bool:
        """Check if email is valid."""
        if not email or '@' not in email or '.' not in email:
            return False
        
        # Basic email validation
        parts = email.split('@')
        if len(parts) != 2:
            return False
        
        local, domain = parts
        if not local or not domain:
            return False
        
        return True
    
    def _looks_like_business_name(self, name: str) -> bool:
        """Check if name looks like a business name."""
        if not name:
            return False
        
        # Business name indicators
        business_indicators = [
            'inc', 'llc', 'corp', 'corporation', 'company', 'co', 'ltd', 'limited',
            'associates', 'assoc', 'group', 'enterprises', 'enterprise', 'services',
            'service', 'solutions', 'solution', 'systems', 'system', 'technologies',
            'technology', 'partners', 'partner', 'ventures', 'venture'
        ]
        
        name_lower = name.lower()
        return any(indicator in name_lower for indicator in business_indicators)
    
    def _clean_contact_export_data(self, data: pd.DataFrame) -> pd.DataFrame:
        """
        Apply Contact Export-specific data cleaning
        
        Args:
            data: Data to clean
            
        Returns:
            Cleaned data
        """
        df_cleaned = data.copy()
        
        # Clean phone numbers
        if 'Phone number' in df_cleaned.columns:
            df_cleaned['Phone number'] = df_cleaned['Phone number'].apply(self._clean_phone)
        
        # Clean state entries (remove leading/trailing spaces)
        if 'State' in df_cleaned.columns:
            df_cleaned['State'] = df_cleaned['State'].apply(lambda x: str(x).strip() if pd.notna(x) else x)
        
        # Clean business type entries
        if 'Business type' in df_cleaned.columns:
            df_cleaned['Business type'] = df_cleaned['Business type'].apply(lambda x: str(x).strip() if pd.notna(x) else x)
        
        return df_cleaned
    
    def _clean_phone(self, phone):
        """Clean phone numbers - return formatted string or None"""
        if pd.isna(phone) or phone == '' or phone == '0':
            return None
        
        phone_str = str(phone).strip()
        
        # If it's just "0", return None
        if phone_str == '0':
            return None
        
        # If all zeros, return None
        if phone_str == '0000000000':
            return None
        
        # Handle float values (from Excel conversion)
        if isinstance(phone, float):
            phone_str = str(int(phone))
        
        # Extract digits only
        import re
        digits = re.sub(r'\D', '', phone_str)
        
        # Format phone number
        if len(digits) == 10:
            return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
        elif len(digits) == 11 and digits[0] == '1':
            return f"1 ({digits[1:4]}) {digits[4:7]}-{digits[7:]}"
        else:
            return None
    
    def create_excel_output(self, processed_data: pd.DataFrame, removed_rows: pd.DataFrame, output_file: str):
        """
        Create Excel output with Contact Export-specific formatting
        
        Args:
            processed_data: Processed data to output
            removed_rows: Rows that were removed during processing
            output_file: Path to output file
        """
        
        wb = Workbook()
        
        # Create the main output sheet
        ws = wb.active
        ws.title = "Contact Export"
        
        # Define headers in the exact order specified
        headers = [
            'Email', 'Business type', 'First name', 'Last name', 'Customer name',
            'Phone number', 'Sales person', 'Address 1', 'Address 2', 'City', 'State', 'Zip'
        ]
        
        # Add headers with formatting
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Map processed data to the correct columns
        for row_idx, (_, row_data) in enumerate(processed_data.iterrows(), 2):
            # Email
            email = row_data.get('email', '')
            ws.cell(row=row_idx, column=1, value=email)
            
            # Business type
            business_type = row_data.get('Business Type', '')
            ws.cell(row=row_idx, column=2, value=business_type)
            
            # First name
            first_name = row_data.get('first_name', '')
            ws.cell(row=row_idx, column=3, value=first_name)
            
            # Last name
            last_name = row_data.get('last_name', '')
            ws.cell(row=row_idx, column=4, value=last_name)
            
            # Customer name
            customer_name = row_data.get('customer_name', '')
            ws.cell(row=row_idx, column=5, value=customer_name)
            
            # Phone number - apply custom formatting
            phone = row_data.get('phone', '')
            if phone and pd.notna(phone):
                # Clean and format phone number
                phone_cleaned = self._clean_phone(phone)
                ws.cell(row=row_idx, column=6, value=phone_cleaned)
            else:
                ws.cell(row=row_idx, column=6, value='')
            
            # Sales person
            salesperson = row_data.get('Salesperson', '')
            ws.cell(row=row_idx, column=7, value=salesperson)
            
            # Address 1
            address_1 = row_data.get('Address 1', '')
            ws.cell(row=row_idx, column=8, value=address_1)
            
            # Address 2
            address_2 = row_data.get('Address 2', '')
            ws.cell(row=row_idx, column=9, value=address_2)
            
            # City
            city = row_data.get('City', '')
            ws.cell(row=row_idx, column=10, value=city)
            
            # State - remove leading spaces
            state = row_data.get('State', '')
            if state and pd.notna(state):
                state_cleaned = str(state).strip()
                ws.cell(row=row_idx, column=11, value=state_cleaned)
            else:
                ws.cell(row=row_idx, column=11, value='')
            
            # Zip
            zip_code = row_data.get('Zip', '')
            ws.cell(row=row_idx, column=12, value=zip_code)
        
        # Create Removed sheet
        if len(removed_rows) > 0:
             ws_removed = wb.create_sheet("Removed")
             
             # Add headers with Removal Reason column at the end
             headers_with_reason = self.original_columns + ['Removal Reason']
             for col_idx, col_name in enumerate(headers_with_reason, 1):
                 cell = ws_removed.cell(row=1, column=col_idx, value=col_name)
                 cell.font = Font(bold=True)
                 cell.alignment = Alignment(horizontal='center')
             
             # Add removed data with removal reasons
             for row_idx, (_, row_data) in enumerate(removed_rows.iterrows(), 2):
                 # Add original data columns first
                 for col_idx, col_name in enumerate(self.original_columns, 1):
                     value = row_data.get(col_name, '')
                     ws_removed.cell(row=row_idx, column=col_idx, value=value)
                 
                 # Add removal reason at the end
                 removal_reason = row_data.get('Removal Reason', 'unknown')
                 ws_removed.cell(row=row_idx, column=len(self.original_columns) + 1, value=removal_reason)
        
        # Save the workbook
        try:
            wb.save(output_file)
            print(f"✅ Excel output created: {output_file}")
            print(f"   Sheets: Contact Export ({len(processed_data)} rows), Removed ({len(removed_rows)} rows)")
        except Exception as e:
            raise RuntimeError(f"Failed to write output file {output_file}: {e}")
    
    def get_processing_stats(self) -> Dict[str, Any]:
        """Get processing statistics from unified processor"""
        return {
            'workflow_stats': {
                'workflow_type': 'contact_export',
                'processor_version': '1.0.0'
            },
            'core_stats': self.data_processor.get_processing_stats(),
            'llm_stats': self.llm_engine.get_stats() if hasattr(self.llm_engine, 'get_stats') else {}
        }
