import pandas as pd
import sys
import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from typing import Optional, Tuple, Dict, Any
import requests
import json

# LLM Configuration
LLM_ENABLED = True  # Set to False to use rule-based only
LLM_MAX_ROWS = 1000  # Process in batches for files with more than this many rows
LLM_BATCH_SIZE = 500  # Process LLM in batches of this size
LLM_MODEL = "qwen2.5:7b-instruct-q4_K_M"  # Clean model without contamination issues
OLLAMA_BASE_URL = "http://localhost:11434"  # Ollama API endpoint
MAX_TOKENS = 150
TEMPERATURE = 0.1

def setup_llm():
    """Setup Ollama client"""
    try:
        # Test connection to Ollama with shorter timeout
        response = requests.get(f"{OLLAMA_BASE_URL}/api/tags", timeout=3)
        if response.status_code == 200:
            return True
        else:
            print(f"⚠️  Ollama connection failed: {response.status_code}")
            return False
    except requests.exceptions.RequestException as e:
        print(f"⚠️  Cannot connect to Ollama at {OLLAMA_BASE_URL}")
        print(f"   Make sure Ollama is running and accessible")
        return False

def reset_llm_state():
    """Reset LLM failure state to enable LLM processing"""
    global LLM_WORKING
    LLM_WORKING = True
    if hasattr(llm_parse_name, '_llm_failed'):
        delattr(llm_parse_name, '_llm_failed')

# Global flag to track if LLM is working
LLM_WORKING = True  # Reset to True to enable LLM

def llm_parse_name(contact_name: str, customer_name: str, email: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Use Ollama LLM to intelligently parse names from Contact Name, Customer Name, and Email
    Returns (first_name, last_name) tuple
    """
    global LLM_WORKING
    
    if not LLM_ENABLED:
        return None, None
    
    # If LLM failed before, don't try again
    if not LLM_WORKING and hasattr(llm_parse_name, '_llm_failed'):
        print(f"   ⚠️  LLM disabled due to previous failure")
        return None, None
    
    if not setup_llm():
        llm_parse_name._llm_failed = True
        return None, None
    
    # Prepare the prompt
    prompt = f"""You are a data cleaning expert. Extract the best person's name from the provided fields.

CRITICAL: Return ONLY valid JSON. NO comments, explanations, or additional text.

ANALYSIS PROCESS:
1. CHECK Customer Name for "LASTNAME, FIRSTNAME" format - this takes HIGHEST priority if it's a person name
2. EXAMINE Contact Name - if it's a clear person name (not company), use it as PRIMARY source
3. ANALYZE Email(s) - extract person names from email addresses ONLY when Contact Name is missing or incomplete
4. COMPARE all sources to find the most complete and accurate person name
5. NEVER use company name fragments as person names
6. PRIORITY ORDER: Customer Name (if person format) > Contact Name > Email extraction

CUSTOMER NAME PRIORITY:
- "LASTNAME, FIRSTNAME" format: Use this when it contains person names
- "SAW, NAHA" becomes first="Naha", last="Saw" 
- "HAN, VICTOR" becomes first="Victor", last="Han" (prioritize over Contact Name)
- "SUBRAMANIAM, SUBBU" format: Use this over incomplete Contact Names like "SUBBU Y"
- If Customer Name is a business, still extract person names from Contact Name + Email

CONTACT NAME FORMATS:
- "LASTNAME, FIRSTNAME": "ERICKSON, MELISSA" becomes first="Melissa", last="Erickson"
- "FIRSTNAME LASTNAME": "JOHN SMITH" becomes first="John", last="Smith"
- "FIRSTNAME LASTNAME - LASTNAME2": "SUZANNE EWING - ERVIN" becomes first="Suzanne", last="Ewing-Ervin"
- "MULTIPLE PEOPLE & SEPARATOR": "KOREY KOENIG-DAMIENS & TAMERA" - extract the most relevant person name
  * If email matches one person (e.g., "tgarlock@berkeley.edu" matches "TAMERA"), use that person
  * "KOREY KOENIG-DAMIENS & TAMERA" + "tgarlock@berkeley.edu" = first="Tamera", last="Garlock"
  * "JOHN & JANE SMITH" + "jsmith@email.com" = first="Jane", last="Smith"
- Always preserve the EXACT names from Contact Name when they are person names
- "SUSAN PARISH" + "sparish@berkeley.edu" = first="Susan", last="Parish" (use Contact Name, not email)
- "MIRIAM" + "mholland@sccoe.org" = first="Miriam", last="Holland" (complete from email when Contact Name incomplete)
- "STACY ROCK" = first="Stacy", last="Rock" (clear person name, even with business email)

EMAIL PARSING GUIDELINES:
- "mholland@domain.com" + Contact Name "MIRIAM": Miriam=first name, Holland=last name (complete the name)
- "SHEILA.SALENGA@domain.com" + Contact Name "SHEILA": Sheila=first name, Salenga=last name
- "sandy@domain.com" + business Customer Name: Sandy=first name (extract from email when no Contact Name)
- "Isela.CornejoEspinoza@domain.com": Isela=first name, Cornejo-Espinoza=last name (preserve hyphenation)
- "tgarlock@berkeley.edu": t=first initial, garlock=last name (extract "Garlock" as last name)
- "Lastname, Firstname <email@domain.com>": Extract "Firstname Lastname" as the person's name
- "Name <email@domain.com>": Extract names from the prefix part before < >
- "MBHOLLAND100@domain.com": M=first initial, B=middle initial, HOLLAND=last name
- "john.smith@domain.com": john=first name, smith=last name  
- "jkolander@company.com": j=first initial, kolander=last name
- COMPLETE NAMES: Use email to fill missing first or last names from Contact Name

MULTIPLE EMAIL HANDLING:
- "sandy@satoricellars.com; tom@satoricellars.com": Choose "sandy@" as it's more personal-sounding
- "SHEILA.SALENGA@LATTICE.COM; workplace-support@lattice.com": Choose the personal name email
- "tgarlock@berkeley.edu;moving@berkeley.edu": Choose "tgarlock@" as it's more personal-sounding
- Choose the email most likely to be a person (not info@, sales@, workplace-support@, etc.)
- Personal emails (firstname.lastname@, firstinitiallastname@) over business emails
- Match email patterns with Contact Name when possible

COMPANY NAME DETECTION:
- ALWAYS extract person names from Contact Name + Email, even when Customer Name is a business
- "BLANCA" + "blanca.ortiz0@wdc.com" = "Blanca Ortiz" (extract last name from email)
- "JACKIE" + "JKOLANDER@company.com" = "Jackie Kolander" (extract last name from email)
- NEVER extract person names from clear business names like "SRI ANANDA BHAVAN RESTAURANT" 
- NEVER use business fragments as person names: "Capital", "Modern", "Senior Living", "Elementary", "Solutions", "Systems", "Research Corporation", "Print Solutions", "Design Decor", "Southern Kitchen", "Climate Week", "Real-Time Innovations", "De Anza", "Maggiore Tile", "Senior Living", "Streamlinevents", "Javier Events", "Cloud Kitchen", "West Gymnastics", "Deli Cafe", "Southern Creations"
- AGGRESSIVE BUSINESS DETECTION: Company indicators include: LLC, INC, CORP, GROUP, CENTER, CATERING, ENTERTAINMENT, RESTAURANT, BHAVAN, SCHOOL, CHURCH, UNIVERSITY, COLLEGE, HOSPITAL, CLINIC, FOUNDATION, ASSOCIATION, ORGANIZATION, INSTITUTE, SYSTEM, NETWORK, PARTNERS, ALLIANCE, PRODUCTIONS, STEAKHOUSE, WINE BAR, INNOVATION, TECHNOLOGY, REAL ESTATE, BOOSTER CLUB, TRADITION, RESIDENCE, DISTRICT, COUNTY, OFFICE, DEPARTMENT, DIVISION, BUREAU, AGENCY, AUTHORITY, COMMISSION, BOARD, COUNCIL, COMMITTEE, TEAM, STAFF, PERSONNEL, EMPLOYEES, WORKERS, MEMBERS, VOLUNTEERS, CAPITAL, MODERN, SENIOR, LIVING, ELEMENTARY, MIDDLE, HIGH, SOLUTIONS, SYSTEMS, INNOVATIONS, RESEARCH, CORPORATION, PRINT, DESIGN, DECOR, TILE, KITCHEN, CREATIONS, STREAMLINE, EXPECTATIONS, SUCCULENT, CACTUS, CLIMATE, WEEK, ZOO, HOTEL, CAFE, DELI, PLAZA, GYMNASTICS, GOVERNMENT, CITY, ARENA, AMERICAS, SOUTHERN, CLOUD, WEST, PACIFIC, LUXE, REAL-TIME, etc.
- NEVER extract these terms as person names: "School", "Church", "University", "Hospital", "Foundation", "Association", "Organization", "Institute", "System", "Network", "Partners", "Alliance", "Productions", "Restaurant", "Steakhouse", "Wine Bar", "Innovation", "Technology", "Real Estate", "Booster Club", "Tradition", "Residence", "District", "County", "Office", "Department", "Division", "Bureau", "Agency", "Authority", "Commission", "Board", "Council", "Committee", "Team", "Staff", "Personnel", "Employees", "Workers", "Members", "Volunteers", "Capital", "Modern", "Senior", "Living", "Elementary", "Solutions", "Systems", "Research", "Corporation", "Print", "Design", "Decor", "Kitchen", "Creations", "Climate", "Week", "Hotel", "Cafe", "Deli", "Gymnastics", "Government", "City", "Southern", "Cloud", "West", "Pacific"
- If Customer Name is business AND Contact Name is empty AND email has no clear person name, return null
- Business email patterns like "SABSV1111@gmail.com" (random letters/numbers) suggest no person name
- ONLY return null when ALL sources (Contact Name, Customer Name, Email) are clearly business-related
- Examples of business names that should return null: "SRI ANANDA BHAVAN RESTAURANT" (no Contact Name), "MANTRA INDIA" (Contact Name is business), "SABSV1111@gmail.com" (business abbreviation email with no Contact Name)
- If Contact Name is a clear person name (like "STACY ROCK"), extract it regardless of business email

HYPHENATED NAMES:
- "EWING - ERVIN" becomes "Ewing-Ervin" (remove spaces around hyphen)
- "CornejoEspinoza" from email becomes "Cornejo-Espinoza" (add hyphen for compound names)
- "MARTINEZ-GARCIA" stays "Martinez-Garcia"
- Always preserve hyphenated last names as single unit

MISSING INFORMATION:
- If only first name available and no reliable last name can be extracted, leave last name null
- Better to have incomplete but accurate data than incorrect data
- For pure business names with no person information, return null for both names

Contact Name: "{contact_name}"
Customer Name: "{customer_name}"
Email: "{email}"

EXAMPLES:
{{"first_name": "Mercedes", "last_name": "Holland"}}
{{"first_name": "Melissa", "last_name": "Erickson"}}
{{"first_name": "Naha", "last_name": "Saw"}}
{{"first_name": "Miriam", "last_name": "Holland"}}
{{"first_name": "Sheila", "last_name": "Salenga"}}
{{"first_name": "Jackie", "last_name": "Kolander"}}
{{"first_name": "Victor", "last_name": "Han"}}
{{"first_name": "Sandy", "last_name": null}}
{{"first_name": "Stacy", "last_name": "Rock"}}
{{"first_name": "Sridhar", "last_name": "Lakshmikanthan"}}
{{"first_name": "Tamera", "last_name": "Garlock"}}

RESPONSE FORMAT: Return ONLY the JSON object. No explanations, comments, or reasoning."""

    try:
        # Call Ollama API with explicit context reset
        response = requests.post(
            f"{OLLAMA_BASE_URL}/api/generate",
            json={
                "model": LLM_MODEL,
                "prompt": prompt,
                "stream": False,
                "context": [],  # Reset context to prevent contamination
                "options": {
                    "temperature": TEMPERATURE,
                    "num_predict": MAX_TOKENS
                }
            },
            timeout=60  # Increased timeout for better reliability
        )
        
        if response.status_code != 200:
            print(f"⚠️  Ollama API error: {response.status_code}")
            llm_parse_name._llm_failed = True
            return None, None
        
        result = response.json()
        content = result.get('response', '').strip()
        
        # Mark LLM as working if we get here
        LLM_WORKING = True
        
        # Parse JSON response with improved error handling
        try:
            # Try to clean up common JSON formatting issues
            content_clean = content.strip()
            
            # Remove any trailing comments or explanations after the JSON
            if '//' in content_clean:
                # Find the first // and remove everything after it, then try to close JSON
                comment_pos = content_clean.find('//')
                json_part = content_clean[:comment_pos].strip()
                # Check if we need to close brackets/braces
                if json_part.endswith(','):
                    json_part = json_part[:-1]  # Remove trailing comma
                if json_part.count('{') > json_part.count('}'):
                    json_part += '}'
                content_clean = json_part
            
            # Extract just the JSON object if there's extra text
            start_brace = content_clean.find('{')
            end_brace = content_clean.rfind('}')
            if start_brace != -1 and end_brace != -1 and end_brace > start_brace:
                content_clean = content_clean[start_brace:end_brace + 1]
            
            parsed = json.loads(content_clean)
            first_name = parsed.get('first_name')
            last_name = parsed.get('last_name')
            
            # Clean up the names
            if first_name and first_name != "null" and first_name != "None":
                first_name = clean_name_for_display(first_name)
            else:
                first_name = None
                
            if last_name and last_name != "null" and last_name != "None":
                last_name = clean_name_for_display(last_name)
                # Validate last name
                if not is_valid_last_name(last_name):
                    last_name = None
            else:
                last_name = None
            
            # Print success message with the parsed names
            print(f"   ✅ LLM successfully processed: {first_name}, {last_name}")
            
            return first_name, last_name
            
        except json.JSONDecodeError:
            llm_parse_name._llm_failed = True
            return None, None
            
    except Exception as e:
        llm_parse_name._llm_failed = True
        return None, None

def clean_name_for_display(name):
    """Clean name for display, handling special characters and formatting"""
    if pd.isna(name) or name == '' or name == "null" or name == "None":
        return None
    
    name_str = str(name).strip()
    
    # Handle hyphenated names with spaces around hyphens
    # "EWING - ERVIN" becomes "Ewing-Ervin"
    name_str = re.sub(r'\s*-\s*', '-', name_str)
    
    # Remove numbers from names (but keep hyphens and other separators)
    name_str = re.sub(r'\d+', '', name_str)
    
    # Remove extra spaces but preserve hyphens
    name_str = ' '.join(name_str.split())
    name_str = name_str.title()
    
    return name_str

def is_valid_email(email):
    """Check if email is valid and not generic"""
    if pd.isna(email) or email == '':
        return False
    
    email_str = str(email).strip()
    
    # Basic email validation
    if '@' not in email_str or '.' not in email_str:
        return False
    
    # Check for common non-email content
    if email_str.lower() in ['fff', 'nan', 'none', 'null']:
        return False
    
    # Check for emails with obvious typos or invalid patterns
    if any(pattern in email_str.lower() for pattern in ['fff', 'nan', 'null', 'none', 'error', 'test']):
        return False
    
    # Check for generic emails - but be more lenient
    if '@' in email_str:
        local_part = email_str.split('@')[0].lower()
        # Only reject very generic emails, not business-specific ones like 'ap'
        if local_part in ['info', 'sales', 'admin', 'contact', 'support', 'help']:
            return False
        
        # Check for suspicious patterns in local part
        if len(local_part) < 2 or local_part.isdigit():
            return False
    
    return True

def clean_email(email, row=None):
    """Clean and validate email addresses, return the best single email"""
    if pd.isna(email) or email == '':
        return None
        
    email_str = str(email).strip()
    
    # Handle "Name <email@domain.com>" format
    if '<' in email_str and '>' in email_str:
        start = email_str.find('<')
        end = email_str.find('>')
        if start != -1 and end != -1 and end > start:
            email_str = email_str[start + 1:end].strip()
    
    # Basic validation
    if not is_valid_email(email_str):
        return None
    
    # Handle multiple emails separated by semicolon
    if ';' in email_str or ',' in email_str:
        # Split by both semicolon and comma
        separators = [';', ',']
        emails = [email_str]
        for sep in separators:
            new_emails = []
            for email in emails:
                new_emails.extend([e.strip() for e in email.split(sep) if e.strip()])
            emails = new_emails
        
        # Filter valid emails
        valid_emails = [e for e in emails if is_valid_email(e)]
        
        if not valid_emails:
            return None
        
        # If we have multiple valid emails, try to find the one that matches the person's name
        if row is not None:
            contact_name = row.get('Contact Name')
            if contact_name and not pd.isna(contact_name):
                contact_str = str(contact_name).strip()
                # Look for email that contains the contact name
                for email in valid_emails:
                    if '@' in email:
                        local_part = email.split('@')[0].lower()
                        if contact_str.lower() in local_part or local_part in contact_str.lower():
                            return email
        
        # Return the first valid email if no match found
        return valid_emails[0]
    
    return email_str

def extract_name_from_email(email):
    """Extract name from email address with improved logic"""
    if pd.isna(email) or email == '':
        return None, None
    
    email_str = str(email).strip()
    
    # Extract the part before @
    if '@' in email_str:
        local_part = email_str.split('@')[0]
        
        # Handle common patterns
        if '.' in local_part:
            # Split by dots and convert to proper case
            parts = local_part.split('.')
            if len(parts) >= 2:
                # Check if parts look like reasonable names (not numbers, not too short)
                if len(parts[0]) >= 2 and len(parts[1]) >= 2 and not parts[0].isdigit() and not parts[1].isdigit():
                    first_name = parts[0].title()
                    last_name = parts[1].title()
                    # Remove numbers from both parts
                    first_name = re.sub(r'\d+', '', first_name)
                    last_name = re.sub(r'\d+', '', last_name)
                    # Only return if neither part is a business term and last name is valid
                    if not is_company_name(first_name) and not is_company_name(last_name) and is_valid_last_name(last_name):
                        return first_name, last_name
        
        # If no dots, try to split by other separators
        for separator in ['_', '-']:
            if separator in local_part:
                parts = local_part.split(separator)
                if len(parts) >= 2:
                    # Check if parts look like reasonable names
                    if len(parts[0]) >= 2 and len(parts[1]) >= 2 and not parts[0].isdigit() and not parts[1].isdigit():
                        first_name = parts[0].title()
                        last_name = parts[1].title()
                        # Remove numbers from both parts
                        first_name = re.sub(r'\d+', '', first_name)
                        last_name = re.sub(r'\d+', '', last_name)
                        # Only return if neither part is a business term and last name is valid
                        if not is_company_name(first_name) and not is_company_name(last_name) and is_valid_last_name(last_name):
                            return first_name, last_name
        
        # Special handling for patterns like "tgarlock" (first initial + last name)
        if len(local_part) >= 4:
            # Check if it starts with a single letter followed by a longer part
            if len(local_part) > 3 and local_part[0].isalpha() and local_part[1:].isalpha():
                # Pattern like "tgarlock" -> "t" + "garlock"
                first_initial = local_part[0].upper()
                last_name_part = local_part[1:].title()
                
                # Remove numbers from last name part
                last_name_part = re.sub(r'\d+', '', last_name_part)
                
                # Only return if the last name part is reasonable (at least 3 characters)
                # and not a business term and is a valid last name
                if len(last_name_part) >= 3 and not is_company_name(last_name_part) and is_valid_last_name(last_name_part):
                    return first_initial, last_name_part
        
        # If no separators found, try to extract a reasonable first name
        # Look for patterns like "firstname123" or "firstname"
        if len(local_part) >= 3:
            # Remove numbers from the end
            clean_part = re.sub(r'\d+$', '', local_part)
            if len(clean_part) >= 2:
                clean_part = re.sub(r'\d+', '', clean_part)  # Remove any remaining numbers
                # Only return if it's not a business term and is a valid first name
                if not is_company_name(clean_part) and len(clean_part) >= 2:
                    return clean_part.title(), None
    
    return None, None

def clean_contact_name(name):
    """Clean contact names and convert to proper case"""
    if pd.isna(name) or name == '':
        return None
    
    name_str = str(name).strip()
    
    # Remove phone numbers and extensions
    name_str = re.sub(r'\s*\d{3}[-.]?\d{3}[-.]?\d{4}', '', name_str)
    name_str = re.sub(r'\s*ext\.?\s*\d+', '', name_str, flags=re.IGNORECASE)
    
    # Convert to proper case
    name_str = name_str.title()
    
    return name_str

def split_name(full_name):
    """Split full name into first and last name - conservative approach like example"""
    if pd.isna(full_name) or full_name == '':
        return None, None
    
    name_str = str(full_name).strip()
    
    # Handle "Last, First" format
    if ',' in name_str:
        parts = name_str.split(',', 1)
        last_name = parts[0].strip().title()
        first_name = parts[1].strip().title() if len(parts) > 1 else ''
        return first_name, last_name
    
    # Handle "First Last" format
    name_parts = name_str.split()
    
    if len(name_parts) == 0:
        return None, None
    elif len(name_parts) == 1:
        # Single name - treat as first name only (like example)
        return name_parts[0].title(), None
    elif len(name_parts) == 2:
        return name_parts[0].title(), name_parts[1].title()
    else:
        # Handle middle initial and multi-part last names
        first_name = name_parts[0].title()
        
        # Check for middle initial (single character, possibly with period)
        if len(name_parts) >= 3 and (len(name_parts[1]) == 1 or (len(name_parts[1]) == 2 and name_parts[1].endswith('.'))):
            # Middle initial found - add to first name
            middle_initial = name_parts[1].upper()
            if not middle_initial.endswith('.'):
                middle_initial += '.'
            first_name = f"{first_name} {middle_initial}"
            last_name = ' '.join([part.title() for part in name_parts[2:]])
        else:
            # Handle hyphenated last names (like "EWING - ERVIN")
            remaining_parts = name_parts[1:]
            last_name_parts = []
            
            for i, part in enumerate(remaining_parts):
                if part == '-':
                    # If this is a hyphen, check if it's part of a hyphenated name
                    if i > 0 and i < len(remaining_parts) - 1:
                        # This is a hyphen between two parts of a hyphenated name
                        # Don't skip it, include it in the last name
                        last_name_parts.append(part)
                    else:
                        # This is a standalone hyphen, skip it
                        continue
                else:
                    last_name_parts.append(part.title())
            
            last_name = ' '.join(last_name_parts)
        
        return first_name, last_name

def extract_name_from_field(field_value):
    """Extract first and last name from a field value"""
    if pd.isna(field_value) or field_value == '':
        return None, None
    
    field_str = str(field_value).strip()
    
    # Check if it's an email
    if '@' in field_str:
        return None, None  # We'll handle email separately
    
    # Use the main company detection function instead of a separate list
    # This ensures consistency across all functions
    
    # Special handling for multiple people separated by "&" or similar
    if ' & ' in field_str or ' AND ' in field_str.upper():
        # Split by "&" or "AND" and try to extract person names
        if ' & ' in field_str:
            parts = field_str.split(' & ')
        else:
            parts = field_str.upper().split(' AND ')
            # Convert back to original case for the parts
            original_parts = field_str.split(' AND ')
            parts = original_parts if len(original_parts) == len(parts) else parts
        
        # Try to extract names from each part
        for part in parts:
            part = part.strip()
            if part and not is_company_name(part):
                # Check if this part looks like a person name
                cleaned_part = clean_name_for_display(part)
                if cleaned_part:
                    first_name, last_name = split_name(cleaned_part)
                    if first_name and len(first_name) >= 2:
                        # Filter out invalid last names
                        if last_name and (len(last_name) <= 1 or last_name in ['.', '-', '_'] or not is_valid_last_name(last_name)):
                            last_name = None
                        return first_name, last_name
    
    # Special handling for mixed person/company names
    # Look for patterns like "PERSON NAME - COMPANY NAME" or "PERSON NAME COMPANY NAME"
    if ' - ' in field_str:
        # Split by " - " and try to extract person name from the first part
        parts = field_str.split(' - ', 1)
        person_part = parts[0].strip()
        company_part = parts[1].strip() if len(parts) > 1 else ""
        
        # Check if the person part contains hyphens (like "EWING - ERVIN")
        # OR if the company part starts with a person name followed by company indicators
        if '-' in person_part and not is_company_name(person_part):
            # This might be a hyphenated last name, so include more of the original string
            # Look for the last occurrence of a company indicator in the full string
            field_upper = field_str.upper()
            last_company_pos = -1
            company_indicators = get_company_indicators()
            for indicator in company_indicators:
                pos = field_upper.rfind(indicator)
                if pos > last_company_pos:
                    last_company_pos = pos
            
            if last_company_pos > 0:
                # Extract everything before the last company indicator
                person_part = field_str[:last_company_pos].strip()
                # Remove trailing hyphens, dashes, or spaces
                person_part = person_part.rstrip(' -')
        
        # Also check if the company part starts with what looks like a person name
        # For example: "ERVIN MORNINGSTAR ENTERTAINMENT" -> "ERVIN" might be part of the person name
        elif company_part and not is_company_name(person_part):
            # Check if the company part starts with a word that could be a person name
            company_words = company_part.split()
            if len(company_words) > 0:
                first_company_word = company_words[0].upper()
                # If the first word of company part looks like a person name (not a company indicator)
                # and the person part is short, try combining them
                company_indicators = get_company_indicators()
                if (first_company_word not in company_indicators and 
                    len(first_company_word) >= 2 and 
                    len(person_part.split()) <= 2):
                    # Try combining person_part with the first word of company_part
                    combined_person = f"{person_part} - {first_company_word}"
                    if not is_company_name(combined_person):
                        person_part = combined_person
        
        # If the person part looks like a person name (not a company), use it
        if not is_company_name(person_part) and len(person_part.split()) >= 2:
            cleaned_name = clean_name_for_display(person_part)
            if cleaned_name:
                first_name, last_name = split_name(cleaned_name)
                if last_name and (len(last_name) <= 1 or last_name in ['.', '-', '_']):
                    last_name = None
                return first_name, last_name
    
    # Check if it looks like a company name
    if is_company_name(field_str):
        return None, None
    
    # Clean and split the name
    cleaned_name = clean_name_for_display(field_str)
    if cleaned_name:
        first_name, last_name = split_name(cleaned_name)
        
        # Filter out invalid last names (like single characters, dots, etc.)
        if last_name and (len(last_name) <= 1 or last_name in ['.', '-', '_'] or not is_valid_last_name(last_name)):
            last_name = None
        
        return first_name, last_name
    
    return None, None

def clean_name_for_display(name):
    """Clean name for display, handling special characters and formatting"""
    if pd.isna(name) or name == '':
        return None
    
    name_str = str(name).strip()
    
    # Handle special characters and encoding issues
    # Replace common problematic characters
    name_str = name_str.replace('â€™', "'")  # Fix apostrophe encoding
    name_str = name_str.replace('â€œ', '"')  # Fix quote encoding
    name_str = name_str.replace('â€', '"')   # Fix quote encoding
    name_str = name_str.replace('Oâ€™Reilly', "O'Reilly")  # Fix specific name
    
    # Remove numbers from names (but keep hyphens and other separators)
    # This handles cases like "Javiercastellanos2" -> "Javiercastellanos"
    name_str = re.sub(r'\d+', '', name_str)
    
    # Remove extra spaces and convert to proper case
    name_str = ' '.join(name_str.split())
    name_str = name_str.title()
    
    return name_str

def is_valid_last_name(last_name):
    """Check if a last name is valid and not a business term"""
    if pd.isna(last_name) or last_name == '':
        return False
    
    last_name_str = str(last_name).strip().upper()
    
    # List of terms that should never be used as last names
    invalid_last_names = [
        'SENIOR', 'LIVING', 'BARNES', 'CERVANTES', 'JAVIER', 'EVEENTS', 'MODERN',
        'CLIMATE', 'WEEK', 'RESEARCH', 'CORPORATION', 'PRINT', 'SOLUTIONS', 'ROSE',
        'ENTERTAINMENT', 'SOUTHERN', 'KITCHEN', 'PSEFTEAS', 'OR', 'HEATHER', 'VON',
        'BORSTEL', 'MOON', 'SECURITY', 'CASTRO', 'CLOUD', 'WEST', 'GYMNASTICS',
        'OF', 'MILPITAS', 'ELEMENTARY', 'MONDAY', 'INC', 'CHANG', 'VENTS', 'DESIGN',
        'DECOR', 'AMERICAS', 'ANZA', 'TILE', 'ESSAGERARDI', 'SIMICH', 'SAMBANOVA',
        'SYSTEMS', 'DE', 'ANZA', 'MAGGIORE', 'X', 'DELI', 'CAFE', 'GREENBERG',
        'CREATIONS', 'REAL-TIME', 'INNOVATIONS', 'STREAMLINEVENTS', 'JCACTUSSUCCULENT',
        'PLURGECATERING', 'ANDERSON', 'SUBRAMANIAM', 'EXPECTATIONS', 'R&J', 'CONSTRUCTION',
        'ISIONARYEVENTSANDDESIGNS', 'SOLIEMANI', 'DESIGN', 'CERVANTES', 'JAVIER',
        'EVEENTS', 'PSEFTEAS', 'OR', 'HEATHER', 'VON', 'BORSTEL', 'MOON', 'SECURITY',
        'CASTRO', 'OF', 'MONDAY', 'VENTS', 'ESSAGERARDI', 'SIMICH', 'X', 'DELI',
        'CAFE', 'GREENBERG', 'ANDERSON', 'SUBRAMANIAM', 'R&J', 'ISIONARYEVENTSANDDESIGNS',
        'SOLIEMANI', 'DESIGN', 'LIVING', 'WEEK', 'KITCHEN', 'GYMNASTICS', 'ENTERTAINMENT',
        'DESIGN', 'SOLUTIONS', 'SYSTEMS', 'CORPORATION', 'RESEARCH', 'INNOVATIONS',
        'REAL-TIME', 'CLOUD', 'WEST', 'SOUTHERN', 'PACIFIC', 'AMERICAS', 'ARENA',
        'HOTEL', 'CAFE', 'DELI', 'PLAZA', 'ZOO', 'STEAKHOUSE', 'WINE', 'BAR',
        'RESTAURANT', 'CATERING', 'VENUE', 'EVENTS', 'CENTER', 'SERVICES'
    ]
    
    # Check if the last name is in the invalid list
    if last_name_str in invalid_last_names:
        return False
    
    # Check if it contains business indicators
    if is_company_name(last_name_str):
        return False
    
    # Check for suspicious patterns
    if any(pattern in last_name_str for pattern in ['/SECURITY', 'OF ', 'ON ', 'X-', 'CAFE', 'DELI']):
        return False
    
    # Check if it's too short (likely not a real name)
    if len(last_name_str) <= 1:
        return False
    
    # Check if it contains only special characters
    if re.match(r'^[^A-Z]+$', last_name_str):
        return False
    
    return True

def get_company_indicators():
    """Get the comprehensive list of company indicators"""
    return [
        # Core business terms
        'GROUP', 'EVENTS', 'CENTER', 'RENTAL', 'SERVICES', 'COMPANY', 'CORP', 'LLC', 'INC',
        'MARSHALL', 'WALK-THRU', 'BAY VIEW', 'EVENT CENTER', 'COAST EVENTS', 'FIRE',
        'GOOGLE', 'YAHOO', 'GMAIL', 'RENTALS', 'EVENT', 'CATERING', 'VENUE', 'BARTENDING',
        'BABES', 'CLUBHOUSE', 'BILL WILSON', 'PASSION FOR', 'IDEAS EVENTS', 'GT4 EVENTS',
        'CONSTRUCTION', 'DISTRICT', 'GOLF', 'WATER', 'COUNTY', 'PALO ALTO HILLS',
        'EXTERNAL RELATIONS', 'HOSPITAL', 'CELLARS', 'UCB', 'EL CAMINO', 'SATORI',
        'UNIVERSITY', 'HEALTH', 'MEDICAL', 'CLINIC', 'FOUNDATION', 'ASSOCIATION',
        'ORGANIZATION', 'INSTITUTE', 'SYSTEM', 'NETWORK', 'PARTNERS', 'ALLIANCE',
        'PRODUCTIONS', 'RESTAURANT', 'STEAKHOUSE', 'WINE BAR', 'INNOVATION', 'TECHNOLOGY',
        'REAL ESTATE', 'BOOSTER CLUB', 'TRADITION', 'ROTORCRAFT', 'PYRAMID', 'RESIDENCE',
        'SCHOOL', 'CHURCH', 'COLLEGE', 'ACADEMY', 'INSTITUTION', 'DEPARTMENT', 'DIVISION',
        'BUREAU', 'AGENCY', 'AUTHORITY', 'COMMISSION', 'BOARD', 'COUNCIL', 'COMMITTEE',
        'TEAM', 'STAFF', 'PERSONNEL', 'EMPLOYEES', 'WORKERS', 'MEMBERS', 'VOLUNTEERS',
        'OFFICE', 'MINISTRY', 'PARISH', 'DIOCESE', 'CONGREGATION', 'SYNAGOGUE', 'TEMPLE',
        'MOSQUE', 'CATHEDRAL', 'CHAPEL', 'SANCTUARY', 'ORATORY', 'BASILICA', 'ABBEY',
        'MONASTERY', 'CONVENT', 'SEMINARY', 'THEOLOGICAL', 'BIBLE', 'RELIGIOUS',
        'SPIRITUAL', 'FAITH', 'WORSHIP', 'PRAYER', 'SERVICE', 'MASS', 'SERMON',
        'MINISTRY', 'OUTREACH', 'MISSION', 'EVANGELISM', 'DISCIPLESHIP', 'FELLOWSHIP',
        'COMMUNITY', 'SOCIETY', 'CLUB', 'LEAGUE', 'FEDERATION', 'UNION', 'GUILD',
        'SORORITY', 'FRATERNITY', 'LODGE', 'ORDER', 'BROTHERHOOD', 'SISTERHOOD',
        'MANTRA', 'INDIA', 'BHAVAN', 'ANANDA', 'SRI', 'RESTAURANT', 'CATERED', 'TOO',
        'TECH', 'TECH.', 'CHILL', 'PICNIC',
        
        # Additional terms from error analysis
        'CAPITAL', 'MODERN', 'SENIOR', 'LIVING', 'ELEMENTARY', 'MIDDLE', 'HIGH',
        'SOLUTIONS', 'SYSTEMS', 'INNOVATIONS', 'RESEARCH', 'CORPORATION', 'CORP.',
        'PRINT', 'DESIGN', 'DECOR', 'TILE', 'KITCHEN', 'CREATIONS', 'ESTHETIKA',
        'STREAMLINE', 'STREAMLINEVENTS', 'EXPECTATIONS', 'SUCCULENT', 'CACTUS',
        'CLIMATE', 'WEEK', 'ZOO', 'STEAKHOUSE', 'HOTEL', 'CAFE', 'DELI', 'PLAZA',
        'GYMNASTICS', 'GOVERNMENT', 'CITY', 'MILPITAS', 'CLARI', 'INTERO', 'FREMONT',
        'ELLORE', 'AESTHETIKA', 'SAMBANOVA', 'DELLA', 'MAGGIORE', 'RYZEN', 'DELUXE',
        'ARENA', 'AMERICAS', 'ANZA', 'MORTON', "MORTON'S", 'SPLURGE', 'SOUTHERN',
        'PLURGECATERING', 'JCACTUSSUCCULENT', 'JAVIER', 'EVEENTS', 'KAPOR', 'COMPASS',
        'EUREST', 'IRVINE', 'COMNS', 'ATLAS', 'BELLA', 'ROSE', 'ENTERTAINMENT',
        'CHERIE', "CHERIE'S", 'ORTHODOX', 'CHURCH', 'ALMADEN', 'SWIM', 'RACKET',
        'EPICUREAN', 'INTUIT', 'MYLAPORECLOUDKITCHEN', 'MYLAPORE', 'CLOUD', 'PACWEST',
        'PACIFIC', 'WEST', 'GARDENIA', 'GATOS', 'AUGUSTINE', 'LUXELEAL', 'DEANZA',
        'THESSAGROUP', 'LDRY', 'HOTELDEANZA', 'GROCA', 'BAYFC', 'REAL-TIME', 'RTI',
        'STANFORD', 'HIGHEXPECTATIONSONLINE',
        
        # Additional problematic terms from user's list
        'BARNES', 'CERVANTES', 'JAVIER', 'EVEENTS', 'PSEFTEAS', 'OR', 'HEATHER', 
        'VON', 'BORSTEL', 'MOON', 'SECURITY', 'CASTRO', 'OF', 'MONDAY', 'VENTS',
        'ESSAGERARDI', 'SIMICH', 'X', 'DELI', 'CAFE', 'GREENBERG', 'ANDERSON',
        'SUBRAMANIAM', 'R&J', 'ISIONARYEVENTSANDDESIGNS', 'SOLIEMANI', 'DESIGN',
        'CERVANTES', 'JAVIER', 'EVEENTS', 'PSEFTEAS', 'OR', 'HEATHER', 'VON', 
        'BORSTEL', 'MOON', 'SECURITY', 'CASTRO', 'OF', 'MONDAY', 'VENTS',
        'ESSAGERARDI', 'SIMICH', 'X', 'DELI', 'CAFE', 'GREENBERG', 'ANDERSON',
        'SUBRAMANIAM', 'R&J', 'ISIONARYEVENTSANDDESIGNS', 'SOLIEMANI', 'DESIGN'
    ]

def is_company_name(name):
    """Check if a name looks like a company name rather than a person's name"""
    if pd.isna(name) or name == '':
        return False
    
    name_str = str(name).strip().upper()
    
    # Get company indicators from the shared function
    company_indicators = get_company_indicators()
    
    # Check if name contains company indicators
    for indicator in company_indicators:
        # Use word boundaries to avoid matching substrings
        if f" {indicator} " in f" {name_str} " or name_str.startswith(f"{indicator} ") or name_str.endswith(f" {indicator}"):
            return True
    
    # Check if name has too many words (likely company) - but be more lenient with "&" separators
    # For hyphenated names, count meaningful words (not just separators)
    words = [word for word in name_str.split() if word not in ['-', '_', '.']]
    
    # If the name contains "&" or "AND", it might be multiple people rather than a company
    if ' & ' in name_str or ' AND ' in name_str:
        # For names with "&" or "AND", allow up to 6 words before considering it a company
        if len(words) > 6:
            return True
    else:
        # For regular names, use the original 3-word limit
        if len(words) > 3:
            return True
    
    # Check if name contains numbers (likely company)
    if any(char.isdigit() for char in name_str):
        return True
    
    # Check for common company patterns
    if 'THE ' in name_str and len(name_str.split()) > 2:
        return True
    
    # Additional checks for common business patterns
    # Check for words that are commonly business-related
    business_words = ['LIVING', 'WEEK', 'KITCHEN', 'GYMNASTICS', 'ENTERTAINMENT', 
                     'DESIGN', 'SOLUTIONS', 'SYSTEMS', 'CORPORATION', 'RESEARCH',
                     'INNOVATIONS', 'REAL-TIME', 'CLOUD', 'WEST', 'SOUTHERN',
                     'PACIFIC', 'AMERICAS', 'ARENA', 'HOTEL', 'CAFE', 'DELI',
                     'PLAZA', 'ZOO', 'STEAKHOUSE', 'WINE', 'BAR', 'RESTAURANT',
                     'CATERING', 'VENUE', 'EVENTS', 'CENTER', 'SERVICES']
    
    for word in business_words:
        if word in name_str:
            return True
    
    # Check for patterns that suggest business names
    if any(pattern in name_str for pattern in ['/SECURITY', 'OF ', 'ON ', 'X-', 'CAFE', 'DELI']):
        return True
    
    # Check if name contains common business abbreviations
    if any(abbr in name_str for abbr in ['INC', 'LLC', 'CORP', 'LTD', 'CO', '& CO']):
        return True
    
    return False

def clean_phone(phone):
    """Clean phone numbers - return just digits for formatting later"""
    if pd.isna(phone) or phone == '':
        return None
    
    phone_str = str(phone).strip()
    
    # If it's just "0", return None
    if phone_str == '0':
        return None
    
    # If all zeros, return None
    if phone_str == '0000000000':
        return None
    
    # Handle float values (from Excel conversion)
    if isinstance(phone, float):
        phone_str = str(int(phone))
    
    # Extract digits only
    digits = re.sub(r'\D', '', phone_str)
    
    # Return digits as string if we have valid phone number
    if len(digits) == 10 or (len(digits) == 11 and digits[0] == '1'):
        return digits  # Return as string, not float
    else:
        return None

def clean_data(input_file):
    """Main data cleaning function"""
    # Reset LLM state at the start of processing
    reset_llm_state()
    
    print(f"📁 Reading input file: {input_file}")
    sys.stdout.flush()  # Force immediate output
    
    # Read the data starting from row 3 (header row)
    df = pd.read_excel(input_file, sheet_name='slsp', header=2)
    
    print(f"📊 Original shape: {df.shape}")
    print(f"🔎 Columns: {list(df.columns)}")
    
    # Create a copy for cleaning
    df_cleaned = df.copy()
    
    # Track removed rows for the "removed" sheet
    removed_rows = []
    
    # 1. Remove rows with "Accounts Receivable" as Salesperson
    initial_count = len(df_cleaned)
    removed_accounts_df = df_cleaned[df_cleaned['Salesperson'].str.strip() == 'ACCOUNTS RECEIVABLE']
    removed_rows.extend(removed_accounts_df.to_dict('records'))
    df_cleaned = df_cleaned[df_cleaned['Salesperson'].str.strip() != 'ACCOUNTS RECEIVABLE']
    removed_accounts = initial_count - len(df_cleaned)
    if removed_accounts > 0:
        print(f"✅ Removed {removed_accounts} rows with 'Accounts Receivable' as Salesperson")
    
    # 2. Clean emails and remove rows with invalid emails
    if 'Email' in df_cleaned.columns:
        df_cleaned['Email'] = df_cleaned.apply(lambda row: clean_email(row['Email'], row), axis=1)
        initial_count = len(df_cleaned)
        invalid_emails_df = df_cleaned[df_cleaned['Email'].isna()]
        removed_rows.extend(invalid_emails_df.to_dict('records'))
        df_cleaned = df_cleaned.dropna(subset=['Email'])
        removed_count = initial_count - len(df_cleaned)
        print(f"✅ Cleaned emails and removed {removed_count} rows with invalid emails")
    
    # Reset index after dropna to ensure contiguous indexing
    df_cleaned = df_cleaned.reset_index(drop=True)
    
    # 3. Extract first and last names using LLM + rule-based approach
    df_cleaned['First Name'] = None
    df_cleaned['Last Name'] = None
    
    llm_count = 0
    rule_count = 0
    
    total_rows = len(df_cleaned)
    print(f"🔄 Processing {total_rows} rows for name extraction...")
    
    # Determine processing strategy
    use_batching = total_rows > LLM_MAX_ROWS and LLM_ENABLED
    if use_batching:
        print(f"📦 Large file detected ({total_rows} rows). Using batch processing with LLM.")
        print(f"   Batch size: {LLM_BATCH_SIZE} rows")
        num_batches = (total_rows + LLM_BATCH_SIZE - 1) // LLM_BATCH_SIZE
        print(f"   Total batches: {num_batches}")
    
    # Process in batches if needed
    if use_batching:
        for batch_num in range(num_batches):
            start_idx = batch_num * LLM_BATCH_SIZE
            end_idx = min((batch_num + 1) * LLM_BATCH_SIZE, total_rows)
            batch_size = end_idx - start_idx
            
            print(f"\n🔄 Processing batch {batch_num + 1}/{num_batches} (rows {start_idx + 1}-{end_idx})...")
            
            # Process this batch with LLM
            for idx in range(start_idx, end_idx):
                row = df_cleaned.iloc[idx]
                
                # Show progress every 10 rows within batch for more frequent updates
                if (idx - start_idx + 1) % 10 == 0:
                    print(f"   Batch progress: {idx - start_idx + 1}/{batch_size} rows...")
                
                # Convert fields to strings for LLM
                contact_name = str(row.get('Contact Name', '')) if not pd.isna(row.get('Contact Name')) else ""
                customer_name = str(row.get('Customer Name', '')) if not pd.isna(row.get('Customer Name')) else ""
                email = str(row.get('Email', '')) if not pd.isna(row.get('Email')) else ""
                
                # Try LLM first
                try:
                    first_name, last_name = llm_parse_name(contact_name, customer_name, email)
                    # Only use LLM result if both first and last names are present
                    if first_name is not None and last_name is not None:
                        df_cleaned.at[idx, 'First Name'] = first_name
                        df_cleaned.at[idx, 'Last Name'] = last_name
                        llm_count += 1
                        continue
                    else:
                        # LLM returned incomplete result - fall back to rule-based
                        if idx == start_idx:  # First row in batch
                            print(f"   ⚠️  LLM returned incomplete result, using rule-based for batch {batch_num + 1}")
                except Exception as e:
                    # If LLM fails, fall back to rule-based
                    if idx == start_idx:  # First row in batch
                        print(f"   ⚠️  LLM failed on first row, switching to rule-based for batch {batch_num + 1}")
                    pass
                
                # Fallback to rule-based approach
                try:
                    # Extract names from all sources
                    contact_first, contact_last = extract_name_from_field(row.get('Contact Name'))
                    customer_first, customer_last = extract_name_from_field(row.get('Customer Name'))
                    
                    # Only extract from email if Contact Name is not empty and not a business name
                    contact_name = row.get('Contact Name')
                    should_extract_from_email = (
                        contact_name and 
                        not pd.isna(contact_name) and 
                        str(contact_name).strip() != '' and
                        not is_company_name(contact_name)
                    )
                    
                    email_first, email_last = extract_name_from_email(row.get('Email')) if (
                        row.get('Email') and 
                        not pd.isna(row.get('Email')) and 
                        should_extract_from_email
                    ) else (None, None)
                    
                    # Build the best possible name by combining information
                    best_first = None
                    best_last = None
                    
                    # Priority for first name: Contact > Customer > Email
                    if contact_first:
                        best_first = contact_first
                    elif customer_first:
                        best_first = customer_first
                    elif email_first:
                        best_first = email_first
                    
                    # Priority for last name: Contact > Customer > Email (but only if Customer is a person name)
                    if contact_last:
                        best_last = contact_last
                    elif customer_last and not is_company_name(row.get('Customer Name')):
                        best_last = customer_last
                    elif email_last:
                        best_last = email_last
                    
                    # Validate last name before assigning
                    if best_last and not is_valid_last_name(best_last):
                        best_last = None
                    
                    # Assign the results
                    df_cleaned.at[idx, 'First Name'] = best_first
                    df_cleaned.at[idx, 'Last Name'] = best_last
                    rule_count += 1
                except Exception as e:
                    print(f"⚠️  Name parsing failed for row {idx}: {e}")
                    rule_count += 1
            
            print(f"✅ Completed batch {batch_num + 1}/{num_batches}")
    else:
        # Process all rows normally (for smaller files)
        for idx, row in df_cleaned.iterrows():
            # Show progress every 50 rows for large files
            if (idx + 1) % 50 == 0:
                print(f"   Processed {idx + 1}/{total_rows} rows... ({(idx + 1)/total_rows*100:.1f}%)")
            
            # Convert fields to strings for LLM
            contact_name = str(row.get('Contact Name', '')) if not pd.isna(row.get('Contact Name')) else ""
            customer_name = str(row.get('Customer Name', '')) if not pd.isna(row.get('Customer Name')) else ""
            email = str(row.get('Email', '')) if not pd.isna(row.get('Email')) else ""
            
            # Try LLM first
            if LLM_ENABLED:
                try:
                    first_name, last_name = llm_parse_name(contact_name, customer_name, email)
                    # Only use LLM result if both first and last names are present
                    if first_name is not None and last_name is not None:
                        df_cleaned.at[idx, 'First Name'] = first_name
                        df_cleaned.at[idx, 'Last Name'] = last_name
                        llm_count += 1
                        continue
                    else:
                        # LLM returned incomplete result - fall back to rule-based
                        pass
                except Exception as e:
                    pass  # Fall back to rule-based
            
            # Fallback to rule-based approach
            try:
                # Extract names from all sources
                contact_first, contact_last = extract_name_from_field(row.get('Contact Name'))
                customer_first, customer_last = extract_name_from_field(row.get('Customer Name'))
                
                # Only extract from email if Contact Name is not empty and not a business name
                contact_name = row.get('Contact Name')
                should_extract_from_email = (
                    contact_name and 
                    not pd.isna(contact_name) and 
                    str(contact_name).strip() != '' and
                    not is_company_name(contact_name)
                )
                
                email_first, email_last = extract_name_from_email(row.get('Email')) if (
                    row.get('Email') and 
                    not pd.isna(row.get('Email')) and 
                    should_extract_from_email
                ) else (None, None)
                
                # Build the best possible name by combining information
                best_first = None
                best_last = None
                
                # Priority for first name: Contact > Customer > Email
                if contact_first:
                    best_first = contact_first
                elif customer_first:
                    best_first = customer_first
                elif email_first:
                    best_first = email_first
                
                # Priority for last name: Contact > Customer > Email (but only if Customer is a person name)
                if contact_last:
                    best_last = contact_last
                elif customer_last and not is_company_name(row.get('Customer Name')):
                    best_last = customer_last
                elif email_last:
                    best_last = email_last
                
                # Validate last name before assigning
                if best_last and not is_valid_last_name(best_last):
                    best_last = None
                
                # Assign the results
                df_cleaned.at[idx, 'First Name'] = best_first
                df_cleaned.at[idx, 'Last Name'] = best_last
                rule_count += 1
            except Exception as e:
                print(f"⚠️  Name parsing failed for row {idx}: {e}")
                rule_count += 1
    
    print(f"✅ Extracted names: {llm_count} LLM-processed, {rule_count} rule-based")
    
    print("🔄 Cleaning additional data...")
    
    # 4. Clean phone numbers
    if 'Phone' in df_cleaned.columns:
        df_cleaned['Phone'] = df_cleaned['Phone'].apply(clean_phone)
        print("✅ Cleaned phone numbers")
    
    # 4.5. Clean State column entries (remove leading/trailing spaces)
    if 'State' in df_cleaned.columns:
        df_cleaned['State'] = df_cleaned['State'].apply(lambda x: str(x).strip() if pd.notna(x) else x)
        print("✅ Cleaned State column entries (removed leading/trailing spaces)")
    
    # 4.6. Clean Business Type entries (use actual Business Type field, not Type field)
    if 'Business Type' in df_cleaned.columns:
        # Just clean the Business Type field without standardization - keep original values
        df_cleaned['Business Type'] = df_cleaned['Business Type'].apply(lambda x: str(x).strip() if pd.notna(x) else x)
        print("✅ Cleaned Business Type entries (removed leading/trailing spaces)")
    print("🔄 Creating output format...")
    
    # 5. Create the output dataframe with the required field order
    output_data = []
    
    for idx, row in df_cleaned.iterrows():
        output_row = {
            'Email': row.get('Email'),
            'Business type': row.get('Business Type', ''),
            'First name': row.get('First Name', ''),
            'Last name': row.get('Last Name', '') if row.get('Last Name') else '',
            'Customer name': row.get('Customer Name', ''),
            'Phone number': row.get('Phone', ''),
            'Sales person': row.get('Salesperson', ''),
            'Address 1': row.get('Address 1', ''),
            'Address 2': row.get('Address 2', '') if pd.notna(row.get('Address 2')) else '',
            'City': row.get('City', ''),
            'State': row.get('State', ''),
            'Zip': row.get('Zip', '')
        }
        output_data.append(output_row)
    
    df_output = pd.DataFrame(output_data)
    
    # 5.5. Remove rows without person names from the output and track them
    initial_count = len(df_output)
    no_person_names_df = df_output[
        ((df_output['First name'].isna()) | (df_output['First name'] == '') | (df_output['First name'] == 'None')) & 
        ((df_output['Last name'].isna()) | (df_output['Last name'] == '') | (df_output['Last name'] == 'None'))
    ]
    
    # Convert the output rows back to original format for the removed sheet
    for idx, row in no_person_names_df.iterrows():
        # Find the corresponding original row by email
        email = row['Email']
        original_row = df_cleaned[df_cleaned['Email'] == email]
        if len(original_row) > 0:
            removed_rows.append(original_row.iloc[0].to_dict())
    
    df_output = df_output[
        ~(((df_output['First name'].isna()) | (df_output['First name'] == '') | (df_output['First name'] == 'None')) & 
          ((df_output['Last name'].isna()) | (df_output['Last name'] == '') | (df_output['Last name'] == 'None')))
    ]
    removed_no_names = initial_count - len(df_output)
    if removed_no_names > 0:
        print(f"✅ Removed {removed_no_names} rows without person names")
    else:
        print("⚠️  No rows were removed (this might indicate an issue)")
    
    # 6. Remove duplicate rows based on case-insensitive email comparison (keep first occurrence)
    initial_count = len(df_output)
    
    # Create a temporary column with lowercase emails for duplicate detection
    df_output['_temp_email_lower'] = df_output['Email'].str.lower()
    df_output = df_output.drop_duplicates(subset=['_temp_email_lower'], keep='first')
    df_output = df_output.drop('_temp_email_lower', axis=1)  # Remove temporary column
    
    removed_count = initial_count - len(df_output)
    if removed_count > 0:
        print(f"✅ Removed {removed_count} duplicate rows based on case-insensitive email")
    
    print(f"📊 Final shape: {df_output.shape}")
    
    return df_output, removed_rows

def create_excel_file(data, filename, removed_rows=None):
    """Create Excel file with proper formatting and removed rows sheet"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Email Contact Export"
    
    # Add headers
    for col_idx, col_name in enumerate(data.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row_idx, row_data in enumerate(data.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            # Handle phone numbers with custom formatting
            if data.columns[col_idx-1] == 'Phone number' and value is not None and value != '':
                try:
                    # Handle both string and float phone numbers
                    if isinstance(value, float):
                        phone_int = int(value)
                    else:
                        phone_int = int(str(value))
                    cell = ws.cell(row=row_idx, column=col_idx, value=phone_int)
                    cell.number_format = '"("* ###")"000"-"0000'  # Custom phone format
                except (ValueError, TypeError):
                    # Fallback to string if conversion fails
                    ws.cell(row=row_idx, column=col_idx, value=str(value))
            else:
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Add removed rows sheet if there are any
    if removed_rows and len(removed_rows) > 0:
        ws_removed = wb.create_sheet("Removed")
        
        # Get the original column names from the first removed row
        if removed_rows:
            original_columns = list(removed_rows[0].keys())
            
            # Add headers to removed sheet
            for col_idx, col_name in enumerate(original_columns, 1):
                cell = ws_removed.cell(row=1, column=col_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Add removed data
            for row_idx, row_data in enumerate(removed_rows, 2):
                for col_idx, col_name in enumerate(original_columns, 1):
                    value = row_data.get(col_name, '')
                    ws_removed.cell(row=row_idx, column=col_idx, value=value)
    
    # Save the workbook
    try:
        wb.save(filename)
    except Exception as e:
        raise RuntimeError(f"Failed to write output file {filename}: {e}")

def main():
    """Main function"""
    try:
        if len(sys.argv) < 2:
            print("❌ Usage: python alert_contact_export_with_llm.py <input_file.xlsx>")
            sys.exit(1)

        input_file = sys.argv[1]
        
        if not os.path.exists(input_file):
            print(f"❌ Input file not found: {input_file}")
            sys.exit(1)
        
        print(f"📁 Input file: {input_file}")
        print(f"LLM Enabled: {LLM_ENABLED}")
        if LLM_ENABLED:
            print(f"LLM Model: {LLM_MODEL}")
            print(f"Ollama URL: {OLLAMA_BASE_URL}")
        
        print("🚀 Starting data cleaning process...")
        sys.stdout.flush()  # Force immediate output
        
        # Clean the data
        cleaned_data, removed_rows = clean_data(input_file)
        
        # Generate output filename and path
        base_name = os.path.splitext(os.path.basename(input_file))[0]
        output_dir = r"C:\LocalAI\Stuart\broadly\Broadly Report\Email Contact Export"
        
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        
        output_filename = os.path.join(output_dir, f"{base_name} - Email Contact Export.xlsx")
        
        # Create the Excel file with removed rows sheet
        create_excel_file(cleaned_data, output_filename, removed_rows)
        
        print("✅ Data cleaning completed successfully!")
        print(f"📄 Output file: {output_filename}")

    except Exception as e:
        print(f"❌ Error occurred: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main() 