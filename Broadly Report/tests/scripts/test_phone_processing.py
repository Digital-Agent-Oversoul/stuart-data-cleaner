import pandas as pd
import sys
import os

def test_phone_processing(input_file):
    """Test phone number processing through the pipeline"""
    print(f"🔍 Testing phone number processing...")
    print(f"📁 Input: {input_file}")
    
    # 1. Read raw data
    print("\n📊 STEP 1: Raw data from Excel")
    df_raw = pd.read_excel(input_file, sheet_name='slsp', header=2)
    print(f"Phone column type: {df_raw['Phone'].dtype}")
    print(f"Phone column sample:")
    for i, phone in enumerate(df_raw['Phone'].head(10)):
        print(f"  {i}: {phone} (type: {type(phone)})")
    
    # 2. Test clean_phone function
    print("\n📊 STEP 2: Testing clean_phone function")
    from alert_contact_export_with_llm import clean_phone
    
    for i, phone in enumerate(df_raw['Phone'].head(10)):
        cleaned = clean_phone(phone)
        print(f"  {i}: {phone} -> {cleaned} (type: {type(cleaned)})")
    
    # 3. Test DataFrame creation
    print("\n📊 STEP 3: Testing DataFrame creation")
    test_data = []
    for i, row in df_raw.head(5).iterrows():
        cleaned_phone = clean_phone(row['Phone'])
        test_data.append({
            'Email': row['Email'],
            'Phone number': cleaned_phone
        })
    
    df_test = pd.DataFrame(test_data)
    print(f"DataFrame phone column type: {df_test['Phone number'].dtype}")
    print(f"DataFrame phone values:")
    for i, phone in enumerate(df_test['Phone number']):
        print(f"  {i}: {phone} (type: {type(phone)})")
    
    # 4. Test Excel writing
    print("\n📊 STEP 4: Testing Excel writing")
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    
    wb = Workbook()
    ws = wb.active
    
    # Add headers
    ws.cell(row=1, column=1, value='Email')
    ws.cell(row=1, column=2, value='Phone number')
    
    # Add data
    for row_idx, row_data in enumerate(df_test.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            if col_idx == 2 and value is not None and value != '':  # Phone column
                phone_str = str(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=phone_str)
                cell.number_format = '("("* ###")"000"-"0000)'
                print(f"  Row {row_idx}: Stored '{phone_str}' with format")
            else:
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save test file
    test_filename = "test_phone_output.xlsx"
    wb.save(test_filename)
    print(f"✅ Saved test file: {test_filename}")
    
    # 5. Read back the test file
    print("\n📊 STEP 5: Reading back test file")
    df_readback = pd.read_excel(test_filename)
    print(f"Readback phone column type: {df_readback['Phone number'].dtype}")
    print(f"Readback phone values:")
    for i, phone in enumerate(df_readback['Phone number']):
        print(f"  {i}: {phone} (type: {type(phone)})")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python test_phone_processing.py <input_file>")
        sys.exit(1)
    
    input_file = sys.argv[1]
    if not os.path.exists(input_file):
        print(f"❌ Input file not found: {input_file}")
        sys.exit(1)
    
    test_phone_processing(input_file) 