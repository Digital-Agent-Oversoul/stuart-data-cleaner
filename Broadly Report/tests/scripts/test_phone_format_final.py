import pandas as pd
import sys
import os
from openpyxl import load_workbook

def test_phone_format_final(output_file):
    """Test that the custom phone format is applied correctly"""
    print(f"🔍 Testing phone format in: {output_file}")
    
    # Read the Excel file with openpyxl to check the actual cell formatting
    wb = load_workbook(output_file)
    ws = wb.active
    
    print(f"Sheet: {ws.title}")
    print(f"Dimensions: {ws.dimensions}")
    
    # Check phone number column
    phone_col_idx = None
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value == 'Phone number':
            phone_col_idx = col_idx
            break
    
    if phone_col_idx is None:
        print("❌ Phone number column not found")
        return
    
    print(f"Phone number column: {phone_col_idx}")
    
    # Check first 10 phone numbers
    print("\n📞 Phone number formatting check:")
    for row_idx in range(2, min(12, ws.max_row + 1)):
        cell = ws.cell(row=row_idx, column=phone_col_idx)
        value = cell.value
        number_format = cell.number_format
        
        print(f"  Row {row_idx}: Value='{value}' Format='{number_format}'")
    
    # Also check with pandas to see the display
    print(f"\n📊 Pandas read (with dtype=str):")
    df = pd.read_excel(output_file, dtype={'Phone number': str})
    phone_sample = df['Phone number'].head(10)
    for i, phone in enumerate(phone_sample):
        print(f"  {i}: {phone}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python test_phone_format_final.py <output_file>")
        sys.exit(1)
    
    output_file = sys.argv[1]
    if not os.path.exists(output_file):
        print(f"❌ Output file not found: {output_file}")
        sys.exit(1)
    
    test_phone_format_final(output_file) 