import pandas as pd
import sys
import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from typing import Optional, Tuple, Dict, Any
import requests
import json

# LLM Configuration
LLM_ENABLED = True  # Set to False to use rule-based only
LLM_MODEL = "qwen2.5:7b-instruct-q4_K_M"  # Clean model without contamination issues
OLLAMA_BASE_URL = "http://localhost:11434"  # Ollama API endpoint
MAX_TOKENS = 150
TEMPERATURE = 0.1

def setup_llm():
    """Setup Ollama client"""
    try:
        # Test connection to Ollama
        response = requests.get(f"{OLLAMA_BASE_URL}/api/tags", timeout=5)
        if response.status_code == 200:
            return True
        else:
            print(f"⚠️  Ollama connection failed: {response.status_code}")
            return False
    except requests.exceptions.RequestException as e:
        print(f"⚠️  Cannot connect to Ollama at {OLLAMA_BASE_URL}")
        print(f"   Make sure Ollama is running and accessible")
        return False

def llm_parse_name(contact_name: str, customer_name: str, email: str) -> Tuple[Optional[str], Optional[str]]:
    """
    Use Ollama LLM to intelligently parse names from Contact Name, Customer Name, and Email
    Returns (first_name, last_name) tuple
    """
    if not LLM_ENABLED:
        return None, None
    
    if not setup_llm():
        return None, None
    
    # Prepare the prompt
    prompt = f"""You are a data cleaning expert. Extract the best person's name from the provided fields.

CRITICAL: Return ONLY valid JSON. NO comments, explanations, or additional text.

ANALYSIS PROCESS:
1. CHECK Customer Name for "LASTNAME, FIRSTNAME" format - this takes HIGHEST priority if it's a person name
2. EXAMINE Contact Name - if it's a clear person name (not company), use it
3. ANALYZE Email(s) - extract person names from email addresses, especially when other fields lack info
4. COMPARE all sources to find the most complete and accurate person name
5. NEVER use company name fragments as person names

CUSTOMER NAME PRIORITY:
- "LASTNAME, FIRSTNAME" format: Use this when it contains person names
- "SAW, NAHA" becomes first="Naha", last="Saw" 
- "HAN, VICTOR" becomes first="Victor", last="Han" (prioritize over Contact Name)
- "SUBRAMANIAM, SUBBU" format: Use this over incomplete Contact Names like "SUBBU Y"
- If Customer Name is a business, still extract person names from Contact Name + Email

CONTACT NAME FORMATS:
- "LASTNAME, FIRSTNAME": "ERICKSON, MELISSA" becomes first="Melissa", last="Erickson"
- "FIRSTNAME LASTNAME": "JOHN SMITH" becomes first="John", last="Smith"
- "FIRSTNAME LASTNAME - LASTNAME2": "SUZANNE EWING - ERVIN" becomes first="Suzanne", last="Ewing-Ervin"
- Always preserve the EXACT names from Contact Name when they are person names

EMAIL PARSING GUIDELINES:
- "mholland@domain.com" + Contact Name "MIRIAM": Miriam=first name, Holland=last name (complete the name)
- "SHEILA.SALENGA@domain.com" + Contact Name "SHEILA": Sheila=first name, Salenga=last name
- "sandy@domain.com" + business Customer Name: Sandy=first name (extract from email when no Contact Name)
- "Isela.CornejoEspinoza@domain.com": Isela=first name, Cornejo-Espinoza=last name (preserve hyphenation)
- "Lastname, Firstname <email@domain.com>": Extract "Firstname Lastname" as the person's name
- "Name <email@domain.com>": Extract names from the prefix part before < >
- "MBHOLLAND100@domain.com": M=first initial, B=middle initial, HOLLAND=last name
- "john.smith@domain.com": john=first name, smith=last name  
- "jkolander@company.com": j=first initial, kolander=last name
- COMPLETE NAMES: Use email to fill missing first or last names from Contact Name

MULTIPLE EMAIL HANDLING:
- "sandy@satoricellars.com; tom@satoricellars.com": Choose "sandy@" as it's more personal-sounding
- "SHEILA.SALENGA@LATTICE.COM; workplace-support@lattice.com": Choose the personal name email
- Choose the email most likely to be a person (not info@, sales@, workplace-support@, etc.)
- Personal emails (firstname.lastname@, firstinitiallastname@) over business emails
- Match email patterns with Contact Name when possible

COMPANY NAME DETECTION:
- ALWAYS extract person names from Contact Name + Email, even when Customer Name is a business
- "BLANCA" + "blanca.ortiz0@wdc.com" = "Blanca Ortiz" (extract last name from email)
- "JACKIE" + "JKOLANDER@company.com" = "Jackie Kolander" (extract last name from email)
- NEVER extract person names from clear business names like "SRI ANANDA BHAVAN RESTAURANT" 
- NEVER use fragments like "Area Asphalt", "India", "And Chill", "Restaurant", "Entertainment" as last names
- Company indicators: LLC, INC, CORP, GROUP, CENTER, CATERING, ENTERTAINMENT, RESTAURANT, BHAVAN, etc.
- If Customer Name is business AND Contact Name is empty AND email has no clear person name, return null
- Business email patterns like "SABSV1111@gmail.com" (random letters/numbers) suggest no person name

HYPHENATED NAMES:
- "EWING - ERVIN" becomes "Ewing-Ervin" (remove spaces around hyphen)
- "CornejoEspinoza" from email becomes "Cornejo-Espinoza" (add hyphen for compound names)
- "MARTINEZ-GARCIA" stays "Martinez-Garcia"
- Always preserve hyphenated last names as single unit

MISSING INFORMATION:
- If only first name available and no reliable last name can be extracted, leave last name null
- Better to have incomplete but accurate data than incorrect data
- For pure business names with no person information, return null for both names

Contact Name: "{contact_name}"
Customer Name: "{customer_name}"
Email: "{email}"

EXAMPLES:
{{"first_name": "Mercedes", "last_name": "Holland"}}
{{"first_name": "Melissa", "last_name": "Erickson"}}
{{"first_name": "Naha", "last_name": "Saw"}}
{{"first_name": "Miriam", "last_name": "Holland"}}
{{"first_name": "Sheila", "last_name": "Salenga"}}
{{"first_name": "Jackie", "last_name": "Kolander"}}
{{"first_name": "Victor", "last_name": "Han"}}
{{"first_name": "Sandy", "last_name": null}}
{{"first_name": "Subbu", "last_name": "Subramaniam"}}
{{"first_name": "Isela", "last_name": "Cornejo-Espinoza"}}
{{"first_name": "Suzanne", "last_name": "Ewing-Ervin"}}
{{"first_name": "Denny", "last_name": "Rahardjo"}}
{{"first_name": "Jocelyn", "last_name": null}}
{{"first_name": null, "last_name": null}}

RESPONSE FORMAT: Return ONLY the JSON object. No explanations, comments, or reasoning."""

    try:
        # Call Ollama API with explicit context reset
        response = requests.post(
            f"{OLLAMA_BASE_URL}/api/generate",
            json={
                "model": LLM_MODEL,
                "prompt": prompt,
                "stream": False,
                "context": [],  # Reset context to prevent contamination
                "options": {
                    "temperature": TEMPERATURE,
                    "num_predict": MAX_TOKENS
                }
            },
            timeout=30
        )
        
        if response.status_code != 200:
            print(f"⚠️  Ollama API error: {response.status_code}")
            return None, None
        
        result = response.json()
        content = result.get('response', '').strip()
        
        # Parse JSON response with improved error handling
        try:
            # Try to clean up common JSON formatting issues
            content_clean = content.strip()
            
            # Remove any trailing comments or explanations after the JSON
            if '//' in content_clean:
                # Find the first // and remove everything after it, then try to close JSON
                comment_pos = content_clean.find('//')
                json_part = content_clean[:comment_pos].strip()
                # Check if we need to close brackets/braces
                if json_part.endswith(','):
                    json_part = json_part[:-1]  # Remove trailing comma
                if json_part.count('{') > json_part.count('}'):
                    json_part += '}'
                content_clean = json_part
            
            # Extract just the JSON object if there's extra text
            start_brace = content_clean.find('{')
            end_brace = content_clean.rfind('}')
            if start_brace != -1 and end_brace != -1 and end_brace > start_brace:
                content_clean = content_clean[start_brace:end_brace + 1]
            
            parsed = json.loads(content_clean)
            first_name = parsed.get('first_name')
            last_name = parsed.get('last_name')
            
            # Clean up the names
            if first_name and first_name != "null":
                first_name = clean_name_for_display(first_name)
            else:
                first_name = None
                
            if last_name and last_name != "null":
                last_name = clean_name_for_display(last_name)
            else:
                last_name = None
            
            return first_name, last_name
            
        except json.JSONDecodeError:
            return None, None
            
    except Exception as e:
        return None, None

def clean_name_for_display(name):
    """Clean name for display, handling special characters and formatting"""
    if pd.isna(name) or name == '' or name == "null":
        return None
    
    name_str = str(name).strip()
    
    # Handle hyphenated names with spaces around hyphens
    # "EWING - ERVIN" becomes "Ewing-Ervin"
    name_str = re.sub(r'\s*-\s*', '-', name_str)
    
    # Remove numbers from names (but keep hyphens and other separators)
    name_str = re.sub(r'\d+', '', name_str)
    
    # Remove extra spaces but preserve hyphens
    name_str = ' '.join(name_str.split())
    name_str = name_str.title()
    
    return name_str

def extract_date_range_from_filename(filename):
    """Extract date range from source filename"""
    # Remove file extension and path
    base_name = os.path.splitext(os.path.basename(filename))[0]
    
    # Look for date patterns like "7.29.25-8.04.25" or "12.04.24 - 01.07.25"
    import re
    
    # Pattern 1: MM.DD.YY-MM.DD.YY
    pattern1 = r'(\d{1,2}\.\d{1,2}\.\d{2})-(\d{1,2}\.\d{1,2}\.\d{2})'
    match1 = re.search(pattern1, base_name)
    if match1:
        return f"{match1.group(1)}-{match1.group(2)}"
    
    # Pattern 2: MM.DD.YY - MM.DD.YY (with spaces)
    pattern2 = r'(\d{1,2}\.\d{1,2}\.\d{2})\s*-\s*(\d{1,2}\.\d{1,2}\.\d{2})'
    match2 = re.search(pattern2, base_name)
    if match2:
        return f"{match2.group(1)}-{match2.group(2)}"
    
    # Pattern 3: MM.DD.YY MM.DD.YY (space separated)
    pattern3 = r'(\d{1,2}\.\d{1,2}\.\d{2})\s+(\d{1,2}\.\d{1,2}\.\d{2})'
    match3 = re.search(pattern3, base_name)
    if match3:
        return f"{match3.group(1)}-{match3.group(2)}"
    
    # If no date range found, use today's date
    today = datetime.now()
    date_str = today.strftime("%m.%d.%Y")
    return date_str

def generate_output_filename(input_filename, location):
    """Generate output filename with date range from source and location"""
    date_range = extract_date_range_from_filename(input_filename)
    return f"{location} - {date_range} - Clean.xlsx"

def is_valid_email(email):
    """Check if email is valid and not generic"""
    if pd.isna(email) or email == '':
        return False
    
    email_str = str(email).strip()
    
    # Basic email validation
    if '@' not in email_str or '.' not in email_str:
        return False
    
    # Check for common non-email content
    if email_str.lower() in ['fff', 'nan', 'none', 'null']:
        return False
    
    # Check for emails with obvious typos or invalid patterns
    if any(pattern in email_str.lower() for pattern in ['fff', 'nan', 'null', 'none', 'error', 'test']):
        return False
    
    # Check for generic emails
    if '@' in email_str:
        local_part = email_str.split('@')[0].lower()
        if local_part in ['info', 'sales', 'ap', 'admin', 'contact', 'support', 'help']:
            return False
        
        # Check for suspicious patterns in local part
        if len(local_part) < 2 or local_part.isdigit():
            return False
    
    return True

def clean_email(email, row=None):
    """Clean and validate email addresses, return the best single email"""
    if pd.isna(email) or email == '':
        return None
        
    email_str = str(email).strip()
    
    # Handle "Name <email@domain.com>" format
    if '<' in email_str and '>' in email_str:
        start = email_str.find('<')
        end = email_str.find('>')
        if start != -1 and end != -1 and end > start:
            email_str = email_str[start + 1:end].strip()
    
    # Basic validation
    if not is_valid_email(email_str):
        return None
    
    # Handle multiple emails separated by semicolon
    if ';' in email_str or ',' in email_str:
        # Split by both semicolon and comma
        separators = [';', ',']
        emails = [email_str]
        for sep in separators:
            new_emails = []
            for email in emails:
                new_emails.extend([e.strip() for e in email.split(sep) if e.strip()])
            emails = new_emails
        
        # Filter valid emails
        valid_emails = [e for e in emails if is_valid_email(e)]
        
        if not valid_emails:
            return None
        
        # If we have multiple valid emails, try to find the one that matches the person's name
        if row is not None:
            contact_name = row.get('Contact Name')
            if contact_name and not pd.isna(contact_name):
                contact_str = str(contact_name).strip()
                # Look for email that contains the contact name
                for email in valid_emails:
                    if '@' in email:
                        local_part = email.split('@')[0].lower()
                        if contact_str.lower() in local_part or local_part in contact_str.lower():
                            return email
        
        # Return the first valid email if no match found
        return valid_emails[0]
    
    return email_str

def extract_name_from_email(email):
    """Extract name from email address with improved logic"""
    if pd.isna(email) or email == '':
        return None, None
    
    email_str = str(email).strip()
    
    # Extract the part before @
    if '@' in email_str:
        local_part = email_str.split('@')[0]
        
        # Handle common patterns
        if '.' in local_part:
            # Split by dots and convert to proper case
            parts = local_part.split('.')
            if len(parts) >= 2:
                # Check if parts look like reasonable names (not numbers, not too short)
                if len(parts[0]) >= 2 and len(parts[1]) >= 2 and not parts[0].isdigit() and not parts[1].isdigit():
                    first_name = parts[0].title()
                    last_name = parts[1].title()
                    # Remove numbers from both parts
                    first_name = re.sub(r'\d+', '', first_name)
                    last_name = re.sub(r'\d+', '', last_name)
                    return first_name, last_name
        
        # If no dots, try to split by other separators
        for separator in ['_', '-']:
            if separator in local_part:
                parts = local_part.split(separator)
                if len(parts) >= 2:
                    # Check if parts look like reasonable names
                    if len(parts[0]) >= 2 and len(parts[1]) >= 2 and not parts[0].isdigit() and not parts[1].isdigit():
                        first_name = parts[0].title()
                        last_name = parts[1].title()
                        # Remove numbers from both parts
                        first_name = re.sub(r'\d+', '', first_name)
                        last_name = re.sub(r'\d+', '', last_name)
                        return first_name, last_name
        
        # If no separators found, try to extract a reasonable first name
        # Look for patterns like "firstname123" or "firstname"
        if len(local_part) >= 3:
            # Remove numbers from the end
            clean_part = re.sub(r'\d+$', '', local_part)
            if len(clean_part) >= 2:
                clean_part = re.sub(r'\d+', '', clean_part)  # Remove any remaining numbers
                return clean_part.title(), None
    
    return None, None

def clean_contact_name(name):
    """Clean contact names and convert to proper case"""
    if pd.isna(name) or name == '':
        return None
    
    name_str = str(name).strip()
    
    # Remove phone numbers and extensions
    name_str = re.sub(r'\s*\d{3}[-.]?\d{3}[-.]?\d{4}', '', name_str)
    name_str = re.sub(r'\s*ext\.?\s*\d+', '', name_str, flags=re.IGNORECASE)
    
    # Convert to proper case
    name_str = name_str.title()
    
    return name_str

def split_name(full_name):
    """Split full name into first and last name - conservative approach like example"""
    if pd.isna(full_name) or full_name == '':
        return None, None
    
    name_str = str(full_name).strip()
    
    # Handle "Last, First" format
    if ',' in name_str:
        parts = name_str.split(',', 1)
        last_name = parts[0].strip().title()
        first_name = parts[1].strip().title() if len(parts) > 1 else ''
        return first_name, last_name
    
    # Handle "First Last" format
    name_parts = name_str.split()
    
    if len(name_parts) == 0:
        return None, None
    elif len(name_parts) == 1:
        # Single name - treat as first name only (like example)
        return name_parts[0].title(), None
    elif len(name_parts) == 2:
        return name_parts[0].title(), name_parts[1].title()
    else:
        # Handle middle initial and multi-part last names
        first_name = name_parts[0].title()
        
        # Check for middle initial (single character, possibly with period)
        if len(name_parts) >= 3 and (len(name_parts[1]) == 1 or (len(name_parts[1]) == 2 and name_parts[1].endswith('.'))):
            # Middle initial found - add to first name
            middle_initial = name_parts[1].upper()
            if not middle_initial.endswith('.'):
                middle_initial += '.'
            first_name = f"{first_name} {middle_initial}"
            last_name = ' '.join([part.title() for part in name_parts[2:]])
        else:
            # Handle hyphenated last names (like "EWING - ERVIN")
            remaining_parts = name_parts[1:]
            last_name_parts = []
            
            for i, part in enumerate(remaining_parts):
                if part == '-':
                    # If this is a hyphen, check if it's part of a hyphenated name
                    if i > 0 and i < len(remaining_parts) - 1:
                        # This is a hyphen between two parts of a hyphenated name
                        # Don't skip it, include it in the last name
                        last_name_parts.append(part)
                    else:
                        # This is a standalone hyphen, skip it
                        continue
                else:
                    last_name_parts.append(part.title())
            
            last_name = ' '.join(last_name_parts)
        
        return first_name, last_name

def determine_best_name(row):
    """Determine the best name by intelligently combining all available sources"""
    contact_name = row.get('Contact Name')
    customer_name = row.get('Customer Name')
    email = row.get('Email')
    
    # Extract names from all sources
    contact_first, contact_last = extract_name_from_field(contact_name)
    customer_first, customer_last = extract_name_from_field(customer_name)
    email_first, email_last = extract_name_from_email(email) if email and not pd.isna(email) else (None, None)
    
    # Build the best possible name by combining information
    best_first = None
    best_last = None
    
    # Priority for first name: Contact > Customer > Email
    if contact_first:
        best_first = contact_first
    elif customer_first:
        best_first = customer_first
    elif email_first:
        best_first = email_first
    
    # Priority for last name: Contact > Customer > Email (but only if Customer is a person name)
    # If Customer is a company name, don't use it for last name
    if contact_last:
        best_last = contact_last
    elif customer_last and not is_company_name(customer_name):
        best_last = customer_last
    elif email_last:
        best_last = email_last
    
    # If we have at least a first name, return the result
    if best_first:
        if best_last:
            return f"{best_first} {best_last}"
        else:
            return best_first
    
    return None

def extract_name_from_field(field_value):
    """Extract first and last name from a field value"""
    if pd.isna(field_value) or field_value == '':
        return None, None
    
    field_str = str(field_value).strip()
    
    # Check if it's an email
    if '@' in field_str:
        return None, None  # We'll handle email separately
    
    # Company indicators for detecting company names
    company_indicators = [
        'GROUP', 'EVENTS', 'CENTER', 'RENTAL', 'SERVICES', 'COMPANY', 'CORP', 'LLC', 'INC',
        'MARSHALL', 'WALK-THRU', 'BAY VIEW', 'EVENT CENTER', 'COAST EVENTS', 'FIRE',
        'GOOGLE', 'YAHOO', 'GMAIL', 'RENTALS', 'EVENT', 'CATERING', 'VENUE', 'BARTENDING',
        'BABES', 'CLUBHOUSE', 'BILL WILSON', 'PASSION FOR', 'IDEAS EVENTS', 'GT4 EVENTS',
        'CONSTRUCTION', 'DISTRICT', 'GOLF', 'WATER', 'COUNTY', 'PALO ALTO HILLS',
        'EXTERNAL RELATIONS', 'HOSPITAL', 'CELLARS', 'UCB', 'EL CAMINO', 'SATORI',
        'UNIVERSITY', 'HEALTH', 'MEDICAL', 'CLINIC', 'FOUNDATION', 'ASSOCIATION',
        'ORGANIZATION', 'INSTITUTE', 'SYSTEM', 'NETWORK', 'PARTNERS', 'ALLIANCE',
        'PRODUCTIONS', 'RESTAURANT', 'STEAKHOUSE', 'WINE BAR', 'INNOVATION', 'TECHNOLOGY',
        'REAL ESTATE', 'BOOSTER CLUB', 'TRADITION', 'ROTORCRAFT', 'PYRAMID', 'RESIDENCE',
        'ENTERTAINMENT'  # Add this specific indicator
    ]
    
    # Special handling for mixed person/company names
    # Look for patterns like "PERSON NAME - COMPANY NAME" or "PERSON NAME COMPANY NAME"
    if ' - ' in field_str:
        # Split by " - " and try to extract person name from the first part
        parts = field_str.split(' - ', 1)
        person_part = parts[0].strip()
        company_part = parts[1].strip() if len(parts) > 1 else ""
        
        # Check if the person part contains hyphens (like "EWING - ERVIN")
        # OR if the company part starts with a person name followed by company indicators
        if '-' in person_part and not is_company_name(person_part):
            # This might be a hyphenated last name, so include more of the original string
            # Look for the last occurrence of a company indicator in the full string
            field_upper = field_str.upper()
            last_company_pos = -1
            for indicator in company_indicators:
                pos = field_upper.rfind(indicator)
                if pos > last_company_pos:
                    last_company_pos = pos
            
            if last_company_pos > 0:
                # Extract everything before the last company indicator
                person_part = field_str[:last_company_pos].strip()
                # Remove trailing hyphens, dashes, or spaces
                person_part = person_part.rstrip(' -')
        
        # Also check if the company part starts with what looks like a person name
        # For example: "ERVIN MORNINGSTAR ENTERTAINMENT" -> "ERVIN" might be part of the person name
        elif company_part and not is_company_name(person_part):
            # Check if the company part starts with a word that could be a person name
            company_words = company_part.split()
            if len(company_words) > 0:
                first_company_word = company_words[0].upper()
                # If the first word of company part looks like a person name (not a company indicator)
                # and the person part is short, try combining them
                if (first_company_word not in company_indicators and 
                    len(first_company_word) >= 2 and 
                    len(person_part.split()) <= 2):
                    # Try combining person_part with the first word of company_part
                    combined_person = f"{person_part} - {first_company_word}"
                    if not is_company_name(combined_person):
                        person_part = combined_person
        
        # If the person part looks like a person name (not a company), use it
        if not is_company_name(person_part) and len(person_part.split()) >= 2:
            cleaned_name = clean_name_for_display(person_part)
            if cleaned_name:
                first_name, last_name = split_name(cleaned_name)
                if last_name and (len(last_name) <= 1 or last_name in ['.', '-', '_']):
                    last_name = None
                return first_name, last_name
    
    # Check if it looks like a company name
    if is_company_name(field_str):
        return None, None
    
    # Clean and split the name
    cleaned_name = clean_name_for_display(field_str)
    if cleaned_name:
        first_name, last_name = split_name(cleaned_name)
        
        # Filter out invalid last names (like single characters, dots, etc.)
        if last_name and (len(last_name) <= 1 or last_name in ['.', '-', '_']):
            last_name = None
        
        return first_name, last_name
    
    return None, None

def clean_name_for_display(name):
    """Clean name for display, handling special characters and formatting"""
    if pd.isna(name) or name == '':
        return None
    
    name_str = str(name).strip()
    
    # Handle special characters and encoding issues
    # Replace common problematic characters
    name_str = name_str.replace('â€™', "'")  # Fix apostrophe encoding
    name_str = name_str.replace('â€œ', '"')  # Fix quote encoding
    name_str = name_str.replace('â€', '"')   # Fix quote encoding
    name_str = name_str.replace('Oâ€™Reilly', "O'Reilly")  # Fix specific name
    
    # Remove numbers from names (but keep hyphens and other separators)
    # This handles cases like "Javiercastellanos2" -> "Javiercastellanos"
    name_str = re.sub(r'\d+', '', name_str)
    
    # Remove extra spaces and convert to proper case
    name_str = ' '.join(name_str.split())
    name_str = name_str.title()
    
    return name_str

def is_company_name(name):
    """Check if a name looks like a company name rather than a person's name"""
    if pd.isna(name) or name == '':
        return False
    
    name_str = str(name).strip().upper()
    
    # Company indicators - be more aggressive
    company_indicators = [
        'GROUP', 'EVENTS', 'CENTER', 'RENTAL', 'SERVICES', 'COMPANY', 'CORP', 'LLC', 'INC',
        'MARSHALL', 'WALK-THRU', 'BAY VIEW', 'EVENT CENTER', 'COAST EVENTS', 'FIRE',
        'GOOGLE', 'YAHOO', 'GMAIL', 'RENTALS', 'EVENT', 'CATERING', 'VENUE', 'BARTENDING',
        'BABES', 'CLUBHOUSE', 'BILL WILSON', 'PASSION FOR', 'IDEAS EVENTS', 'GT4 EVENTS',
        'CONSTRUCTION', 'DISTRICT', 'GOLF', 'WATER', 'COUNTY', 'PALO ALTO HILLS',
        'EXTERNAL RELATIONS', 'HOSPITAL', 'CELLARS', 'UCB', 'EL CAMINO', 'SATORI',
        'UNIVERSITY', 'HEALTH', 'MEDICAL', 'CLINIC', 'FOUNDATION', 'ASSOCIATION',
        'ORGANIZATION', 'INSTITUTE', 'SYSTEM', 'NETWORK', 'PARTNERS', 'ALLIANCE',
        'PRODUCTIONS', 'RESTAURANT', 'STEAKHOUSE', 'WINE BAR', 'INNOVATION', 'TECHNOLOGY',
        'REAL ESTATE', 'BOOSTER CLUB', 'TRADITION', 'ROTORCRAFT', 'PYRAMID', 'RESIDENCE'
    ]
    
    # Check if name contains company indicators
    for indicator in company_indicators:
        if indicator in name_str:
            return True
    
    # Check if name has too many words (likely company)
    # For hyphenated names, count meaningful words (not just separators)
    words = [word for word in name_str.split() if word not in ['-', '_', '.']]
    if len(words) > 3:
        return True
    
    # Check if name contains numbers (likely company)
    if any(char.isdigit() for char in name_str):
        return True
    
    # Check for common company patterns
    if 'THE ' in name_str and len(name_str.split()) > 2:
        return True
    
    return False

def clean_phone(phone):
    """Clean phone numbers - return just digits for formatting later"""
    if pd.isna(phone) or phone == '':
        return None
    
    phone_str = str(phone).strip()
    
    # If it's just "0", return None
    if phone_str == '0':
        return None
    
    # If all zeros, return None
    if phone_str == '0000000000':
        return None
    
    # Extract digits only
    digits = re.sub(r'\D', '', phone_str)
    
    # Return digits if we have valid phone number
    if len(digits) == 10 or (len(digits) == 11 and digits[0] == '1'):
        return digits
    else:
        return None

def determine_location(row):
    """Determine location based on Salesperson and DEL/PU fields"""
    salesperson = str(row.get('Salesperson', '')).strip()
    del_field = str(row.get('DEL', '')).strip()
    pu_field = str(row.get('PU', '')).strip()
    
    # Combine DEL and PU fields for checking
    del_pu_text = f"{del_field} {pu_field}".upper()
    
    # Check for Milpitas override first (highest priority)
    if any(milpitas_indicator in del_pu_text for milpitas_indicator in ['MILPITAS', 'MILPITA', 'MILP']):
        return 'Milpitas'
    
    # Check for Dublin classification
    if (salesperson in ['Mark Pringle', 'Cindy Foster'] or 
        'DUBLIN' in del_pu_text or 'DUB' in del_pu_text):
        return 'Dublin'
    
    # Everything else goes to Milpitas
    return 'Milpitas'

def clean_data(input_file):
    """Main data cleaning function"""
    print(f"📁 Reading input file: {input_file}")
    
    # Read the data starting from row 3 (header row)
    df = pd.read_excel(input_file, sheet_name='slsp', header=2)
    
    print(f"📊 Original shape: {df.shape}")
    print(f"🔎 Columns: {list(df.columns)}")
    
    # Create a copy for cleaning
    df_cleaned = df.copy()
    
    # 1. Remove rows with "Accounts Receivable" as Salesperson
    initial_count = len(df_cleaned)
    df_cleaned = df_cleaned[df_cleaned['Salesperson'].str.strip() != 'ACCOUNTS RECEIVABLE']
    removed_accounts = initial_count - len(df_cleaned)
    if removed_accounts > 0:
        print(f"✅ Removed {removed_accounts} rows with 'Accounts Receivable' as Salesperson")
    
    # 2. Clean emails and remove rows with invalid emails
    if 'Email' in df_cleaned.columns:
        df_cleaned['Email'] = df_cleaned.apply(lambda row: clean_email(row['Email'], row), axis=1)
        initial_count = len(df_cleaned)
        df_cleaned = df_cleaned.dropna(subset=['Email'])
        removed_count = initial_count - len(df_cleaned)
        print(f"✅ Cleaned emails and removed {removed_count} rows with invalid emails")
    
    # 3. Determine best name from available sources
    df_cleaned['Best Name'] = df_cleaned.apply(determine_best_name, axis=1)
    print("✅ Determined best names from available sources")
    
    # 4. Extract first and last names using LLM + rule-based approach
    df_cleaned['First Name'] = None
    df_cleaned['Last Name'] = None
    
    llm_count = 0
    rule_count = 0
    
    for idx, row in df_cleaned.iterrows():
        # Convert fields to strings for LLM
        contact_name = str(row.get('Contact Name', '')) if not pd.isna(row.get('Contact Name')) else ""
        customer_name = str(row.get('Customer Name', '')) if not pd.isna(row.get('Customer Name')) else ""
        email = str(row.get('Email', '')) if not pd.isna(row.get('Email')) else ""
        
        # Try LLM first
        if LLM_ENABLED:
            try:
                first_name, last_name = llm_parse_name(contact_name, customer_name, email)
                # Trust LLM result even if it's None, None (business names)
                df_cleaned.at[idx, 'First Name'] = first_name
                df_cleaned.at[idx, 'Last Name'] = last_name
                llm_count += 1
                continue
            except Exception as e:
                pass  # Fall back to rule-based
        
        # Fallback to rule-based approach
        try:
            # Extract names from all sources
            contact_first, contact_last = extract_name_from_field(row.get('Contact Name'))
            customer_first, customer_last = extract_name_from_field(row.get('Customer Name'))
            email_first, email_last = extract_name_from_email(row.get('Email')) if row.get('Email') and not pd.isna(row.get('Email')) else (None, None)
            
            # Build the best possible name by combining information
            best_first = None
            best_last = None
            
            # Priority for first name: Contact > Customer > Email
            if contact_first:
                best_first = contact_first
            elif customer_first:
                best_first = customer_first
            elif email_first:
                best_first = email_first
            
            # Priority for last name: Contact > Customer > Email (but only if Customer is a person name)
            if contact_last:
                best_last = contact_last
            elif customer_last and not is_company_name(row.get('Customer Name')):
                best_last = customer_last
            elif email_last:
                best_last = email_last
            
            # Assign the results
            df_cleaned.at[idx, 'First Name'] = best_first
            df_cleaned.at[idx, 'Last Name'] = best_last
            rule_count += 1
        except Exception as e:
            print(f"⚠️  Name parsing failed for row {idx}: {e}")
            rule_count += 1
    
    print(f"✅ Extracted names: {llm_count} LLM-processed, {rule_count} rule-based")
    
    # 5. Clean phone numbers
    if 'Phone' in df_cleaned.columns:
        df_cleaned['Phone'] = df_cleaned['Phone'].apply(clean_phone)
        print("✅ Cleaned phone numbers")
    
    # 6. Use original Salesperson as Assisting Team Member
    df_cleaned['Assisting Team Member'] = df_cleaned['Salesperson']
    print("✅ Used original Salesperson as Assisting Team Member")
    
    # 7. Determine location
    df_cleaned['Location'] = df_cleaned.apply(determine_location, axis=1)
    print("✅ Determined locations")
    
    # 8. Select only the required columns
    required_columns = ['Email', 'First Name', 'Last Name', 'Phone', 'Assisting Team Member']
    df_final = df_cleaned[required_columns].copy()
    print("✅ Selected required columns")
    
    # 9. Remove duplicate rows based on email (keep first occurrence)
    initial_count = len(df_final)
    df_final = df_final.drop_duplicates(subset=['Email'], keep='first')
    removed_count = initial_count - len(df_final)
    if removed_count > 0:
        print(f"✅ Removed {removed_count} duplicate rows based on email")
    
    # 10. Sort by Assisting Team Member
    df_final = df_final.sort_values(by=['Assisting Team Member'])
    print("✅ Sorted by Assisting Team Member")
    
    print(f"📊 Final shape: {df_final.shape}")
    
    # 11. Split data by location and create separate files
    # Add location to df_final for filtering
    df_final['Location'] = df_cleaned['Location']
    
    # Filter by location
    dublin_data = df_final[df_final['Location'] == 'Dublin'].copy()
    milpitas_data = df_final[df_final['Location'] == 'Milpitas'].copy()
    
    # Remove the Location column from the output data
    dublin_data = dublin_data.drop('Location', axis=1)
    milpitas_data = milpitas_data.drop('Location', axis=1)
    
    print(f"📍 Dublin records: {len(dublin_data)}")
    print(f"📍 Milpitas records: {len(milpitas_data)}")
    
    # Create Dublin output file
    if len(dublin_data) > 0:
        dublin_filename = generate_output_filename(input_file, 'Dublin')
        create_excel_file(dublin_data, dublin_filename)
        print(f"💾 Dublin data saved to: {dublin_filename}")
    
    # Create Milpitas output file
    if len(milpitas_data) > 0:
        milpitas_filename = generate_output_filename(input_file, 'Milpitas')
        create_excel_file(milpitas_data, milpitas_filename)
        print(f"💾 Milpitas data saved to: {milpitas_filename}")
    
    return dublin_filename if len(dublin_data) > 0 else None, milpitas_filename if len(milpitas_data) > 0 else None

def create_excel_file(data, filename):
    """Create Excel file with proper formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cleaned Data"
    
    # Add headers
    for col_idx, col_name in enumerate(data.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row_idx, row_data in enumerate(data.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            # Handle phone numbers with proper formatting
            if data.columns[col_idx-1] == 'Phone' and value is not None:
                try:
                    # Convert to integer and apply custom formatting
                    phone_int = int(float(value))
                    cell = ws.cell(row=row_idx, column=col_idx, value=phone_int)
                    # Apply custom phone number format (simplified for compatibility)
                    cell.number_format = '(000) 000-0000'
                except (ValueError, TypeError):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            else:
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save the workbook
    try:
        wb.save(filename)
    except Exception as e:
        raise RuntimeError(f"Failed to write output file {filename}: {e}")

def main():
    """Main function"""
    try:
        if len(sys.argv) < 2:
            print("Usage: python broadly_data_cleaner.py <input_file.xlsx>")
            sys.exit(1)

        input_file = sys.argv[1]
        
        if not os.path.exists(input_file):
            print(f"Input file not found: {input_file}")
            sys.exit(1)
        
        print(f"Input file: {input_file}")
        
        # Extract date range for display
        date_range = extract_date_range_from_filename(input_file)
        print(f"Date range detected: {date_range}")
        
        print(f"LLM Enabled: {LLM_ENABLED}")
        if LLM_ENABLED:
            print(f"LLM Model: {LLM_MODEL}")
            print(f"Ollama URL: {OLLAMA_BASE_URL}")

        dublin_file, milpitas_file = clean_data(input_file)
        
        print("Data cleaning completed successfully!")
        if dublin_file:
            print(f"Dublin file: {dublin_file}")
        if milpitas_file:
            print(f"Milpitas file: {milpitas_file}")

    except Exception as e:
        print(f"Error occurred: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main() 