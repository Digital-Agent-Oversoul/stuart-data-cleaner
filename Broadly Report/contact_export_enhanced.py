"""
Enhanced Contact Export Data Cleaner
Uses the unified core LLM engine with OpenAI + Ollama fallback
"""

import pandas as pd
import sys
import os
import re
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from typing import Optional, Tuple, Dict, Any, List
import logging

# Import the unified core engine
from core_llm_engine import create_llm_engine, LLMResponse

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('contact_export.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

class EnhancedContactExportCleaner:
    """Enhanced Contact Export data cleaner using unified LLM engine"""
    
    def __init__(self, config_file: str = "config.json", openai_api_key: Optional[str] = None):
        self.config = self._load_config(config_file)
        self.llm_engine = create_llm_engine(
            openai_api_key=openai_api_key,
            monthly_budget=self.config.get("llm", {}).get("budget", {}).get("monthly_limit", 10.0)
        )
        self.stats = {
            "total_rows": 0,
            "processed_rows": 0,
            "llm_processed": 0,
            "rule_based_fallbacks": 0,
            "uncertain_cases": 0,
            "removed_items": 0,
            "start_time": datetime.now()
        }
        
        # Load configuration
        self.batch_size = self.config.get("processing", {}).get("batch_size", 500)
        self.max_rows_for_llm = self.config.get("processing", {}).get("max_rows_for_llm", 1000)
        self.output_config = self.config.get("output", {}).get("contact_export", {})
        
        logger.info("Enhanced Contact Export Cleaner initialized")
        logger.info(f"LLM Engine: OpenAI {'enabled' if openai_api_key else 'disabled'}, Ollama enabled")
    
    def _load_config(self, config_file: str) -> Dict[str, Any]:
        """Load configuration from JSON file"""
        try:
            if os.path.exists(config_file):
                with open(config_file, 'r') as f:
                    return json.load(f)
            else:
                logger.warning(f"Config file {config_file} not found, using defaults")
                return {}
        except Exception as e:
            logger.error(f"Failed to load config: {e}")
            return {}
    
    def clean_data(self, input_file: str, output_directory: Optional[str] = None) -> bool:
        """Main data cleaning function"""
        try:
            logger.info(f"Starting contact export data cleaning for: {input_file}")
            
            # Read the Excel file
            df = self._read_excel_file(input_file)
            if df is None:
                return False
            
            self.stats["total_rows"] = len(df)
            logger.info(f"Loaded {len(df)} rows from {input_file}")
            
            # Clean the data
            cleaned_df, removed_df = self._clean_dataframe(df)
            
            # Save the cleaned data
            success = self._save_contact_export(cleaned_df, removed_df, output_directory)
            
            if success:
                self._print_summary()
                self._save_processing_stats()
            
            return success
            
        except Exception as e:
            logger.error(f"Data cleaning failed: {e}")
            return False
    
    def _read_excel_file(self, file_path: str) -> Optional[pd.DataFrame]:
        """Read Excel file with proper sheet and header handling"""
        try:
            # Try to read the 'slsp' sheet first
            try:
                df = pd.read_excel(file_path, sheet_name='slsp', header=2)  # Header row 3
                logger.info("Successfully read 'slsp' sheet")
            except Exception as e:
                logger.warning(f"Could not read 'slsp' sheet: {e}")
                # Try to read the first sheet
                df = pd.read_excel(file_path, header=2)
                logger.info("Read first sheet instead")
            
            # Check if we have the expected columns
            expected_columns = ['Contact Name', 'Customer Name', 'Email']
            missing_columns = [col for col in expected_columns if col not in df.columns]
            
            if missing_columns:
                logger.error(f"Missing required columns: {missing_columns}")
                logger.info(f"Available columns: {list(df.columns)}")
                return None
            
            return df
            
        except Exception as e:
            logger.error(f"Failed to read Excel file {file_path}: {e}")
            return None
    
    def _clean_dataframe(self, df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """Clean the dataframe using the unified LLM engine"""
        logger.info("Starting data cleaning process...")
        
        # Create a copy for cleaning
        cleaned_df = df.copy()
        
        # Add new columns for cleaned data
        cleaned_df['First Name'] = None
        cleaned_df['Last Name'] = None
        cleaned_df['Clean Email'] = None
        cleaned_df['Processing Method'] = None
        cleaned_df['Confidence'] = None
        cleaned_df['Notes'] = None
        cleaned_df['Valid Contact'] = True
        
        # Process in batches if file is large
        if len(df) > self.max_rows_for_llm:
            logger.info(f"Large file detected ({len(df)} rows), processing in batches of {self.batch_size}")
            self._process_in_batches(cleaned_df)
        else:
            logger.info("Processing all rows with LLM")
            self._process_all_rows(cleaned_df)
        
        # Separate valid contacts from removed items
        valid_contacts = cleaned_df[cleaned_df['Valid Contact'] == True].copy()
        removed_items = cleaned_df[cleaned_df['Valid Contact'] == False].copy()
        
        self.stats["removed_items"] = len(removed_items)
        
        logger.info(f"Valid contacts: {len(valid_contacts)}, Removed items: {len(removed_items)}")
        
        return valid_contacts, removed_items
    
    def _process_in_batches(self, df: pd.DataFrame):
        """Process data in batches to manage memory and API usage"""
        total_batches = (len(df) + self.batch_size - 1) // self.batch_size
        
        for batch_num in range(total_batches):
            start_idx = batch_num * self.batch_size
            end_idx = min(start_idx + self.batch_size, len(df))
            
            logger.info(f"Processing batch {batch_num + 1}/{total_batches} (rows {start_idx + 1}-{end_idx})")
            
            # Process this batch
            self._process_batch(df, start_idx, end_idx)
            
            # Print progress
            progress = ((batch_num + 1) / total_batches) * 100
            logger.info(f"Progress: {progress:.1f}% complete")
            
            # Check budget status
            budget_status = self.llm_engine.get_stats()["budget_status"]
            if budget_status["percentage_used"] > 80:
                logger.warning(f"⚠️  Budget alert: {budget_status['percentage_used']:.1f}% used")
    
    def _process_batch(self, df: pd.DataFrame, start_idx: int, end_idx: int):
        """Process a batch of rows"""
        for idx in range(start_idx, end_idx):
            if idx % 100 == 0:
                logger.info(f"  Processing row {idx + 1}/{end_idx}")
            
            self._process_single_row(df, idx)
    
    def _process_all_rows(self, df: pd.DataFrame):
        """Process all rows with LLM"""
        for idx in range(len(df)):
            if idx % 100 == 0:
                logger.info(f"Processing row {idx + 1}/{len(df)}")
            
            self._process_single_row(df, idx)
    
    def _process_single_row(self, df: pd.DataFrame, idx: int):
        """Process a single row of data"""
        try:
            row = df.iloc[idx]
            
            # Extract fields
            contact_name = str(row['Contact Name']) if pd.notna(row['Contact Name']) else ""
            customer_name = str(row['Customer Name']) if pd.notna(row['Customer Name']) else ""
            email = str(row['Email']) if pd.notna(row['Email']) else ""
            
            # Clean email
            clean_email = self._clean_email(email)
            
            # Parse names using LLM engine
            first_name, last_name = self.llm_engine.parse_name(contact_name, customer_name, email)
            
            # Determine if this is a valid contact
            is_valid = self._is_valid_contact(first_name, last_name, clean_email, contact_name, customer_name)
            
            # Update the dataframe
            df.at[idx, 'First Name'] = first_name
            df.at[idx, 'Last Name'] = last_name
            df.at[idx, 'Clean Email'] = clean_email
            df.at[idx, 'Processing Method'] = 'LLM'
            df.at[idx, 'Confidence'] = 0.8  # Default confidence for LLM
            df.at[idx, 'Valid Contact'] = is_valid
            
            if not is_valid:
                df.at[idx, 'Notes'] = 'Removed: Invalid or business contact'
            else:
                df.at[idx, 'Notes'] = ''
            
            self.stats["processed_rows"] += 1
            self.stats["llm_processed"] += 1
            
        except Exception as e:
            logger.error(f"Error processing row {idx}: {e}")
            df.at[idx, 'Notes'] = f"Processing error: {str(e)}"
            df.at[idx, 'Valid Contact'] = False
    
    def _clean_email(self, email: str) -> str:
        """Clean and validate email addresses"""
        if not email or pd.isna(email):
            return ""
        
        email = str(email).strip()
        
        # Remove common prefixes and suffixes
        email = re.sub(r'^.*?<(.+?)>.*$', r'\1', email)  # Extract from <email@domain.com>
        email = re.sub(r'^.*?([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}).*$', r'\1', email)
        
        # Basic email validation
        if re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
            return email.lower()
        
        return ""
    
    def _is_valid_contact(self, first_name: Optional[str], last_name: Optional[str], 
                          email: str, contact_name: str, customer_name: str) -> bool:
        """Determine if this is a valid contact for export"""
        # Must have at least a first name or last name
        if not first_name and not last_name:
            return False
        
        # Must have a valid email
        if not email or not re.match(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', email):
            return False
        
        # Check if this looks like a business name rather than a person
        if self._is_business_name(contact_name) or self._is_business_name(customer_name):
            return False
        
        # Check for common business email patterns
        if self._is_business_email(email):
            return False
        
        return True
    
    def _is_business_name(self, name: str) -> bool:
        """Detect if a name is likely a business name"""
        if not name:
            return False
        
        business_indicators = [
            'LLC', 'INC', 'CORP', 'GROUP', 'CENTER', 'RESTAURANT', 'ENTERTAINMENT',
            'BHAVAN', 'CATERING', 'SERVICES', 'SOLUTIONS', 'TECHNOLOGIES', 'SYSTEMS',
            'COMPANY', 'ASSOCIATES', 'PARTNERS', 'ENTERPRISES', 'VENTURES'
        ]
        
        name_upper = name.upper()
        return any(indicator in name_upper for indicator in business_indicators)
    
    def _is_business_email(self, email: str) -> bool:
        """Detect if an email is likely a business email"""
        if not email:
            return False
        
        business_patterns = [
            r'^info@', r'^sales@', r'^support@', r'^admin@', r'^contact@',
            r'^hello@', r'^team@', r'^office@', r'^service@', r'^help@',
            r'^workplace-support@', r'^noreply@', r'^donotreply@'
        ]
        
        return any(re.match(pattern, email.lower()) for pattern in business_patterns)
    
    def _save_contact_export(self, cleaned_df: pd.DataFrame, removed_df: pd.DataFrame, 
                            output_directory: Optional[str] = None) -> bool:
        """Save the cleaned contact export data"""
        try:
            if output_directory is None:
                output_directory = self.output_config.get("base_directory", "Broadly Report/Email Contact Export")
            
            # Ensure output directory exists
            os.makedirs(output_directory, exist_ok=True)
            
            # Generate timestamp for filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            base_filename = self.output_config.get("filename_prefix", "Contact_Export")
            
            # Save main contact export file
            main_filename = f"{base_filename}_{timestamp}.xlsx"
            main_filepath = os.path.join(output_directory, main_filename)
            
            # Create Excel file with multiple sheets
            with pd.ExcelWriter(main_filepath, engine='openpyxl') as writer:
                # Main contacts sheet
                cleaned_df.to_excel(writer, sheet_name='Contacts', index=False)
                
                # Removed items sheet (if any)
                if len(removed_df) > 0:
                    removed_df.to_excel(writer, sheet_name='Removed Items', index=False)
                
                # Get the workbook for formatting
                workbook = writer.book
                
                # Format the Contacts sheet
                self._format_contacts_sheet(workbook['Contacts'])
                
                # Format the Removed Items sheet if it exists
                if 'Removed Items' in workbook.sheetnames:
                    self._format_removed_items_sheet(workbook['Removed Items'])
            
            logger.info(f"Contact export saved: {main_filepath}")
            logger.info(f"Main contacts: {len(cleaned_df)} rows")
            logger.info(f"Removed items: {len(removed_df)} rows")
            
            return True
            
        except Exception as e:
            logger.error(f"Failed to save contact export: {e}")
            return False
    
    def _format_contacts_sheet(self, worksheet):
        """Format the contacts worksheet"""
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Format header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Add filters
        worksheet.auto_filter.ref = worksheet.dimensions
    
    def _format_removed_items_sheet(self, worksheet):
        """Format the removed items worksheet"""
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Format header row with different color
        header_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Add filters
        worksheet.auto_filter.ref = worksheet.dimensions
    
    def _print_summary(self):
        """Print processing summary"""
        processing_time = datetime.now() - self.stats["start_time"]
        
        print("\n" + "="*60)
        print("CONTACT EXPORT DATA CLEANING SUMMARY")
        print("="*60)
        print(f"Total rows processed: {self.stats['total_rows']}")
        print(f"Successfully processed: {self.stats['processed_rows']}")
        print(f"Valid contacts: {self.stats['processed_rows'] - self.stats['removed_items']}")
        print(f"Removed items: {self.stats['removed_items']}")
        print(f"LLM processed: {self.stats['llm_processed']}")
        print(f"Rule-based fallbacks: {self.stats['rule_based_fallbacks']}")
        print(f"Uncertain cases: {self.stats['uncertain_cases']}")
        print(f"Processing time: {processing_time.total_seconds():.2f} seconds")
        
        # Print LLM engine stats
        llm_stats = self.llm_engine.get_stats()
        print(f"\nLLM Engine Statistics:")
        print(f"  Total requests: {llm_stats['total_requests']}")
        print(f"  OpenAI requests: {llm_stats['openai_requests']}")
        print(f"  Ollama requests: {llm_stats['ollama_requests']}")
        print(f"  Rule-based fallbacks: {llm_stats['rule_based_fallbacks']}")
        
        # Print budget status
        budget_status = llm_stats['budget_status']
        print(f"\nBudget Status:")
        print(f"  Current cost: ${budget_status['current_cost']:.4f}")
        print(f"  Monthly budget: ${budget_status['monthly_budget']:.2f}")
        print(f"  Remaining: ${budget_status['remaining_budget']:.2f}")
        print(f"  Percentage used: {budget_status['percentage_used']:.1f}%")
        
        if budget_status['percentage_used'] > 80:
            print(f"  ⚠️  BUDGET ALERT: {budget_status['percentage_used']:.1f}% of budget used!")
        
        print("="*60)
    
    def _save_processing_stats(self):
        """Save processing statistics to file"""
        try:
            stats_file = f"contact_export_stats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            
            stats_data = {
                "timestamp": datetime.now().isoformat(),
                "processing_stats": self.stats,
                "llm_engine_stats": self.llm_engine.get_stats(),
                "config": self.config
            }
            
            with open(stats_file, 'w') as f:
                json.dump(stats_data, f, indent=2)
            
            logger.info(f"Processing statistics saved to: {stats_file}")
            
        except Exception as e:
            logger.error(f"Failed to save processing statistics: {e}")

def main():
    """Main function for command-line usage"""
    import argparse
    
    parser = argparse.ArgumentParser(description="Enhanced Contact Export Data Cleaner")
    parser.add_argument("input_file", help="Path to input Excel file")
    parser.add_argument("--output-dir", help="Output directory (optional)")
    parser.add_argument("--config", default="config.json", help="Configuration file path")
    parser.add_argument("--openai-key", help="OpenAI API key (or set OPENAI_API_KEY env var)")
    
    args = parser.parse_args()
    
    # Get OpenAI API key from arguments or environment
    openai_api_key = args.openai_key or os.getenv("OPENAI_API_KEY")
    
    if not openai_api_key:
        logger.warning("No OpenAI API key provided - OpenAI processing will be disabled")
        logger.info("Set OPENAI_API_KEY environment variable or use --openai-key argument")
    
    # Create cleaner instance
    cleaner = EnhancedContactExportCleaner(
        config_file=args.config,
        openai_api_key=openai_api_key
    )
    
    # Process the file
    success = cleaner.clean_data(args.input_file, args.output_dir)
    
    if success:
        logger.info("Contact export data cleaning completed successfully!")
        sys.exit(0)
    else:
        logger.error("Contact export data cleaning failed!")
        sys.exit(1)

if __name__ == "__main__":
    main()
