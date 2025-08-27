import pandas as pd
import sys
import os
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

def is_valid_email(email):
    """Check if email is valid and not generic"""
    if pd.isna(email) or email == '':
        return False
    
    email_str = str(email).strip()
    
    # Basic email validation
    if '@' not in email_str or '.' not in email_str:
        return False
    
    # Check for common non-email content
    if email_str.lower() in ['fff', 'nan', 'none', 'null']:
        return False
    
    # Check for emails with obvious typos or invalid patterns
    if any(pattern in email_str.lower() for pattern in ['fff', 'nan', 'null', 'none', 'error', 'test']):
        return False
    
    # Check for generic emails
    if '@' in email_str:
        local_part = email_str.split('@')[0].lower()
        if local_part in ['info', 'sales', 'ap', 'admin', 'contact', 'support', 'help']:
            return False
        
        # Check for suspicious patterns in local part
        if len(local_part) < 2 or local_part.isdigit():
            return False
    
    return True

def clean_email(email, row=None):
    """Clean and validate email addresses, return the best single email"""
    if not is_valid_email(email):
        return None
    
    email_str = str(email).strip()
    
    # Handle multiple emails separated by semicolon
    if ';' in email_str or ',' in email_str:
        # Split by both semicolon and comma
        separators = [';', ',']
        emails = [email_str]
        for sep in separators:
            new_emails = []
            for email in emails:
                new_emails.extend([e.strip() for e in email.split(sep) if e.strip()])
            emails = new_emails
        
        # Filter valid emails
        valid_emails = [e for e in emails if is_valid_email(e)]
        
        if not valid_emails:
            return None
        
        # If we have multiple valid emails, try to find the one that matches the person's name
        if row is not None:
            contact_name = row.get('Contact Name')
            if contact_name and not pd.isna(contact_name):
                contact_str = str(contact_name).strip()
                # Look for email that contains the contact name
                for email in valid_emails:
                    if '@' in email:
                        local_part = email.split('@')[0].lower()
                        if contact_str.lower() in local_part or local_part in contact_str.lower():
                            return email
        
        # Return the first valid email if no match found
        return valid_emails[0]
    
    return email_str

def extract_name_from_email(email):
    """Extract name from email address with improved logic"""
    if pd.isna(email) or email == '':
        return None, None
    
    email_str = str(email).strip()
    
    # Extract the part before @
    if '@' in email_str:
        local_part = email_str.split('@')[0]
        
        # Handle common patterns
        if '.' in local_part:
            # Split by dots and convert to proper case
            parts = local_part.split('.')
            if len(parts) >= 2:
                # Check if parts look like reasonable names (not numbers, not too short)
                if len(parts[0]) >= 2 and len(parts[1]) >= 2 and not parts[0].isdigit() and not parts[1].isdigit():
                    first_name = parts[0].title()
                    last_name = parts[1].title()
                    # Remove numbers from both parts
                    first_name = re.sub(r'\d+', '', first_name)
                    last_name = re.sub(r'\d+', '', last_name)
                    return first_name, last_name
        
        # If no dots, try to split by other separators
        for separator in ['_', '-']:
            if separator in local_part:
                parts = local_part.split(separator)
                if len(parts) >= 2:
                    # Check if parts look like reasonable names
                    if len(parts[0]) >= 2 and len(parts[1]) >= 2 and not parts[0].isdigit() and not parts[1].isdigit():
                        first_name = parts[0].title()
                        last_name = parts[1].title()
                        # Remove numbers from both parts
                        first_name = re.sub(r'\d+', '', first_name)
                        last_name = re.sub(r'\d+', '', last_name)
                        return first_name, last_name
        
        # If no separators found, try to extract a reasonable first name
        # Look for patterns like "firstname123" or "firstname"
        if len(local_part) >= 3:
            # Remove numbers from the end
            clean_part = re.sub(r'\d+$', '', local_part)
            if len(clean_part) >= 2:
                clean_part = re.sub(r'\d+', '', clean_part)  # Remove any remaining numbers
                return clean_part.title(), None
    
    return None, None

def clean_contact_name(name):
    """Clean contact names and convert to proper case"""
    if pd.isna(name) or name == '':
        return None
    
    name_str = str(name).strip()
    
    # Remove phone numbers and extensions
    name_str = re.sub(r'\s*\d{3}[-.]?\d{3}[-.]?\d{4}', '', name_str)
    name_str = re.sub(r'\s*ext\.?\s*\d+', '', name_str, flags=re.IGNORECASE)
    
    # Convert to proper case
    name_str = name_str.title()
    
    return name_str

def split_name(full_name):
    """Split full name into first and last name - conservative approach like example"""
    if pd.isna(full_name) or full_name == '':
        return None, None
    
    name_str = str(full_name).strip()
    
    # Handle "Last, First" format
    if ',' in name_str:
        parts = name_str.split(',', 1)
        last_name = parts[0].strip().title()
        first_name = parts[1].strip().title() if len(parts) > 1 else ''
        return first_name, last_name
    
    # Handle "First Last" format
    name_parts = name_str.split()
    
    if len(name_parts) == 0:
        return None, None
    elif len(name_parts) == 1:
        # Single name - treat as first name only (like example)
        return name_parts[0].title(), None
    elif len(name_parts) == 2:
        return name_parts[0].title(), name_parts[1].title()
    else:
        # Handle middle initial and multi-part last names
        first_name = name_parts[0].title()
        
        # Check for middle initial (single character, possibly with period)
        if len(name_parts) >= 3 and (len(name_parts[1]) == 1 or (len(name_parts[1]) == 2 and name_parts[1].endswith('.'))):
            # Middle initial found - add to first name
            middle_initial = name_parts[1].upper()
            if not middle_initial.endswith('.'):
                middle_initial += '.'
            first_name = f"{first_name} {middle_initial}"
            last_name = ' '.join([part.title() for part in name_parts[2:]])
        else:
            # Handle hyphenated last names (like "EWING - ERVIN")
            remaining_parts = name_parts[1:]
            last_name_parts = []
            
            for i, part in enumerate(remaining_parts):
                if part == '-':
                    # If this is a hyphen, check if it's part of a hyphenated name
                    if i > 0 and i < len(remaining_parts) - 1:
                        # This is a hyphen between two parts of a hyphenated name
                        # Don't skip it, include it in the last name
                        last_name_parts.append(part)
                    else:
                        # This is a standalone hyphen, skip it
                        continue
                else:
                    last_name_parts.append(part.title())
            
            last_name = ' '.join(last_name_parts)
        
        return first_name, last_name

def extract_name_from_field(field_value):
    """Extract first and last name from a field value"""
    if pd.isna(field_value) or field_value == '':
        return None, None
    
    field_str = str(field_value).strip()
    
    # Check if it's an email
    if '@' in field_str:
        return None, None  # We'll handle email separately
    
    # Company indicators for detecting company names
    company_indicators = [
        'GROUP', 'EVENTS', 'CENTER', 'RENTAL', 'SERVICES', 'COMPANY', 'CORP', 'LLC', 'INC',
        'MARSHALL', 'WALK-THRU', 'BAY VIEW', 'EVENT CENTER', 'COAST EVENTS', 'FIRE',
        'GOOGLE', 'YAHOO', 'GMAIL', 'RENTALS', 'EVENT', 'CATERING', 'VENUE', 'BARTENDING',
        'BABES', 'CLUBHOUSE', 'BILL WILSON', 'PASSION FOR', 'IDEAS EVENTS', 'GT4 EVENTS',
        'CONSTRUCTION', 'DISTRICT', 'GOLF', 'WATER', 'COUNTY', 'PALO ALTO HILLS',
        'EXTERNAL RELATIONS', 'HOSPITAL', 'CELLARS', 'UCB', 'EL CAMINO', 'SATORI',
        'UNIVERSITY', 'HEALTH', 'MEDICAL', 'CLINIC', 'FOUNDATION', 'ASSOCIATION',
        'ORGANIZATION', 'INSTITUTE', 'SYSTEM', 'NETWORK', 'PARTNERS', 'ALLIANCE',
        'PRODUCTIONS', 'RESTAURANT', 'STEAKHOUSE', 'WINE BAR', 'INNOVATION', 'TECHNOLOGY',
        'REAL ESTATE', 'BOOSTER CLUB', 'TRADITION', 'ROTORCRAFT', 'PYRAMID', 'RESIDENCE',
        'ENTERTAINMENT'  # Add this specific indicator
    ]
    
    # Special handling for mixed person/company names
    # Look for patterns like "PERSON NAME - COMPANY NAME" or "PERSON NAME COMPANY NAME"
    if ' - ' in field_str:
        # Split by " - " and try to extract person name from the first part
        parts = field_str.split(' - ', 1)
        person_part = parts[0].strip()
        company_part = parts[1].strip() if len(parts) > 1 else ""
        
        # Check if the person part contains hyphens (like "EWING - ERVIN")
        # OR if the company part starts with a person name followed by company indicators
        if '-' in person_part and not is_company_name(person_part):
            # This might be a hyphenated last name, so include more of the original string
            # Look for the last occurrence of a company indicator in the full string
            field_upper = field_str.upper()
            last_company_pos = -1
            for indicator in company_indicators:
                pos = field_upper.rfind(indicator)
                if pos > last_company_pos:
                    last_company_pos = pos
            
            if last_company_pos > 0:
                # Extract everything before the last company indicator
                person_part = field_str[:last_company_pos].strip()
                # Remove trailing hyphens, dashes, or spaces
                person_part = person_part.rstrip(' -')
        
        # Also check if the company part starts with what looks like a person name
        # For example: "ERVIN MORNINGSTAR ENTERTAINMENT" -> "ERVIN" might be part of the person name
        elif company_part and not is_company_name(person_part):
            # Check if the company part starts with a word that could be a person name
            company_words = company_part.split()
            if len(company_words) > 0:
                first_company_word = company_words[0].upper()
                # If the first word of company part looks like a person name (not a company indicator)
                # and the person part is short, try combining them
                if (first_company_word not in company_indicators and 
                    len(first_company_word) >= 2 and 
                    len(person_part.split()) <= 2):
                    # Try combining person_part with the first word of company_part
                    combined_person = f"{person_part} - {first_company_word}"
                    if not is_company_name(combined_person):
                        person_part = combined_person
        
        # If the person part looks like a person name (not a company), use it
        if not is_company_name(person_part) and len(person_part.split()) >= 2:
            cleaned_name = clean_name_for_display(person_part)
            if cleaned_name:
                first_name, last_name = split_name(cleaned_name)
                if last_name and (len(last_name) <= 1 or last_name in ['.', '-', '_']):
                    last_name = None
                return first_name, last_name
    
    # Check if it looks like a company name
    if is_company_name(field_str):
        return None, None
    
    # Clean and split the name
    cleaned_name = clean_name_for_display(field_str)
    if cleaned_name:
        first_name, last_name = split_name(cleaned_name)
        
        # Filter out invalid last names (like single characters, dots, etc.)
        if last_name and (len(last_name) <= 1 or last_name in ['.', '-', '_']):
            last_name = None
        
        return first_name, last_name
    
    return None, None

def clean_name_for_display(name):
    """Clean name for display, handling special characters and formatting"""
    if pd.isna(name) or name == '':
        return None
    
    name_str = str(name).strip()
    
    # Handle special characters and encoding issues
    # Replace common problematic characters
    name_str = name_str.replace('â€™', "'")  # Fix apostrophe encoding
    name_str = name_str.replace('â€œ', '"')  # Fix quote encoding
    name_str = name_str.replace('â€', '"')   # Fix quote encoding
    name_str = name_str.replace('Oâ€™Reilly', "O'Reilly")  # Fix specific name
    
    # Remove numbers from names (but keep hyphens and other separators)
    # This handles cases like "Javiercastellanos2" -> "Javiercastellanos"
    name_str = re.sub(r'\d+', '', name_str)
    
    # Remove extra spaces and convert to proper case
    name_str = ' '.join(name_str.split())
    name_str = name_str.title()
    
    return name_str

def is_company_name(name):
    """Check if a name looks like a company name rather than a person's name"""
    if pd.isna(name) or name == '':
        return False
    
    name_str = str(name).strip().upper()
    
    # Company indicators - be more aggressive
    company_indicators = [
        'GROUP', 'EVENTS', 'CENTER', 'RENTAL', 'SERVICES', 'COMPANY', 'CORP', 'LLC', 'INC',
        'MARSHALL', 'WALK-THRU', 'BAY VIEW', 'EVENT CENTER', 'COAST EVENTS', 'FIRE',
        'GOOGLE', 'YAHOO', 'GMAIL', 'RENTALS', 'EVENT', 'CATERING', 'VENUE', 'BARTENDING',
        'BABES', 'CLUBHOUSE', 'BILL WILSON', 'PASSION FOR', 'IDEAS EVENTS', 'GT4 EVENTS',
        'CONSTRUCTION', 'DISTRICT', 'GOLF', 'WATER', 'COUNTY', 'PALO ALTO HILLS',
        'EXTERNAL RELATIONS', 'HOSPITAL', 'CELLARS', 'UCB', 'EL CAMINO', 'SATORI',
        'UNIVERSITY', 'HEALTH', 'MEDICAL', 'CLINIC', 'FOUNDATION', 'ASSOCIATION',
        'ORGANIZATION', 'INSTITUTE', 'SYSTEM', 'NETWORK', 'PARTNERS', 'ALLIANCE',
        'PRODUCTIONS', 'RESTAURANT', 'STEAKHOUSE', 'WINE BAR', 'INNOVATION', 'TECHNOLOGY',
        'REAL ESTATE', 'BOOSTER CLUB', 'TRADITION', 'ROTORCRAFT', 'PYRAMID', 'RESIDENCE'
    ]
    
    # Check if name contains company indicators
    for indicator in company_indicators:
        if indicator in name_str:
            return True
    
    # Check if name has too many words (likely company)
    # For hyphenated names, count meaningful words (not just separators)
    words = [word for word in name_str.split() if word not in ['-', '_', '.']]
    if len(words) > 3:
        return True
    
    # Check if name contains numbers (likely company)
    if any(char.isdigit() for char in name_str):
        return True
    
    # Check for common company patterns
    if 'THE ' in name_str and len(name_str.split()) > 2:
        return True
    
    return False

def clean_phone(phone):
    """Clean phone numbers and return just digits for custom formatting"""
    if pd.isna(phone) or phone == '':
        return None
    
    phone_str = str(phone).strip()
    
    # If it's just "0", return None
    if phone_str == '0':
        return None
    
    # If all zeros, return None
    if phone_str == '0000000000':
        return None
    
    # Extract digits only
    digits = re.sub(r'\D', '', phone_str)
    
    # Return digits if we have valid phone number
    if len(digits) == 10 or (len(digits) == 11 and digits[0] == '1'):
        return digits
    else:
        return None

def determine_best_name(row):
    """Determine the best name by intelligently combining all available sources"""
    contact_name = row.get('Contact Name')
    customer_name = row.get('Customer Name')
    email = row.get('Email')
    
    # Extract names from all sources
    contact_first, contact_last = extract_name_from_field(contact_name)
    customer_first, customer_last = extract_name_from_field(customer_name)
    email_first, email_last = extract_name_from_email(email) if email and not pd.isna(email) else (None, None)
    
    # Build the best possible name by combining information
    best_first = None
    best_last = None
    
    # Priority for first name: Contact > Customer > Email
    if contact_first:
        best_first = contact_first
    elif customer_first:
        best_first = customer_first
    elif email_first:
        best_first = email_first
    
    # Priority for last name: Contact > Customer > Email (but only if Customer is a person name)
    # If Customer is a company name, don't use it for last name
    if contact_last:
        best_last = contact_last
    elif customer_last and not is_company_name(customer_name):
        best_last = customer_last
    elif email_last:
        best_last = email_last
    
    # If we have at least a first name, return the result
    if best_first:
        if best_last:
            return f"{best_first} {best_last}"
        else:
            return best_first
    
    return None

def clean_data(input_file):
    """Main data cleaning function"""
    print(f"📁 Reading input file: {input_file}")
    
    # Read the data starting from row 3 (header row)
    df = pd.read_excel(input_file, sheet_name='slsp', header=2)
    
    print(f"📊 Original shape: {df.shape}")
    print(f"🔎 Columns: {list(df.columns)}")
    
    # Create a copy for cleaning
    df_cleaned = df.copy()
    
    # 1. Remove rows with "Accounts Receivable" as Salesperson
    initial_count = len(df_cleaned)
    df_cleaned = df_cleaned[df_cleaned['Salesperson'].str.strip() != 'ACCOUNTS RECEIVABLE']
    removed_accounts = initial_count - len(df_cleaned)
    if removed_accounts > 0:
        print(f"✅ Removed {removed_accounts} rows with 'Accounts Receivable' as Salesperson")
    
    # 2. Clean emails and remove rows with invalid emails
    if 'Email' in df_cleaned.columns:
        df_cleaned['Email'] = df_cleaned.apply(lambda row: clean_email(row['Email'], row), axis=1)
        initial_count = len(df_cleaned)
        df_cleaned = df_cleaned.dropna(subset=['Email'])
        removed_count = initial_count - len(df_cleaned)
        print(f"✅ Cleaned emails and removed {removed_count} rows with invalid emails")
    
    # 3. Extract first and last names using the same logic as the original script
    df_cleaned['First Name'] = None
    df_cleaned['Last Name'] = None
    
    for idx, row in df_cleaned.iterrows():
        # Use the same intelligent name extraction logic
        contact_name = row.get('Contact Name')
        customer_name = row.get('Customer Name')
        email = row.get('Email')
        
        # Extract names from all sources
        contact_first, contact_last = extract_name_from_field(contact_name)
        customer_first, customer_last = extract_name_from_field(customer_name)
        email_first, email_last = extract_name_from_email(email) if email and not pd.isna(email) else (None, None)
        
        # Build the best possible name by combining information
        best_first = None
        best_last = None
        
        # Priority for first name: Contact > Customer > Email
        if contact_first:
            best_first = contact_first
        elif customer_first:
            best_first = customer_first
        elif email_first:
            best_first = email_first
        
        # Priority for last name: Contact > Customer > Email (but only if Customer is a person name)
        # If Customer is a company name, don't use it for last name
        if contact_last:
            best_last = contact_last
        elif customer_last and not is_company_name(customer_name):
            best_last = customer_last
        elif email_last:
            best_last = email_last
        
        # Assign the results
        df_cleaned.at[idx, 'First Name'] = best_first
        df_cleaned.at[idx, 'Last Name'] = best_last
    
    print("✅ Split names into first and last name")
    
    # 4. Clean phone numbers
    if 'Phone' in df_cleaned.columns:
        df_cleaned['Phone'] = df_cleaned['Phone'].apply(clean_phone)
        print("✅ Cleaned phone numbers")
    
    # 4.5. Clean State column entries (remove leading/trailing spaces)
    if 'State' in df_cleaned.columns:
        df_cleaned['State'] = df_cleaned['State'].apply(lambda x: str(x).strip() if pd.notna(x) else x)
        print("✅ Cleaned State column entries (removed leading/trailing spaces)")
    
    # 4.6. Clean Business Type entries (use actual Business Type field, not Type field)
    if 'Business Type' in df_cleaned.columns:
        # Just clean the Business Type field without standardization - keep original values
        df_cleaned['Business Type'] = df_cleaned['Business Type'].apply(lambda x: str(x).strip() if pd.notna(x) else x)
        print("✅ Cleaned Business Type entries (removed leading/trailing spaces)")
    
    # 5. Create the output dataframe with the required field order
    output_data = []
    
    for idx, row in df_cleaned.iterrows():
        output_row = {
            'Email': row.get('Email'),
            'Business type': row.get('Business Type', ''),
            'First name': row.get('First Name', ''),
            'Last name': row.get('Last Name', '') if row.get('Last Name') else '',
            'Customer name': row.get('Customer Name', ''),
            'Phone number': row.get('Phone', ''),
            'Sales person': row.get('Salesperson', ''),
            'Address 1': row.get('Address 1', ''),
            'Address 2': row.get('Address 2', '') if pd.notna(row.get('Address 2')) else '',
            'City': row.get('City', ''),
            'State': row.get('State', ''),
            'Zip': row.get('Zip', '')
        }
        output_data.append(output_row)
    
    df_output = pd.DataFrame(output_data)
    
    # 6. Remove duplicate rows based on email (keep first occurrence)
    initial_count = len(df_output)
    df_output = df_output.drop_duplicates(subset=['Email'], keep='first')
    removed_count = initial_count - len(df_output)
    if removed_count > 0:
        print(f"✅ Removed {removed_count} duplicate rows based on email")
    
    print(f"📊 Final shape: {df_output.shape}")
    
    return df_output

def create_excel_file(data, filename):
    """Create Excel file with proper formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Alert Contact Export"
    
    # Add headers
    for col_idx, col_name in enumerate(data.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    for row_idx, row_data in enumerate(data.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            # Handle phone numbers with custom formatting
            if data.columns[col_idx-1] == 'Phone number' and value is not None and value != '':
                try:
                    # Convert to integer and apply custom formatting
                    phone_int = int(float(value))
                    cell = ws.cell(row=row_idx, column=col_idx, value=phone_int)
                    # Apply custom phone number format: "("* ###")"000"-"0000
                    cell.number_format = '"("* ###")"000"-"0000'
                except (ValueError, TypeError):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            else:
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Save the workbook
    try:
        wb.save(filename)
    except Exception as e:
        raise RuntimeError(f"Failed to write output file {filename}: {e}")

def main():
    """Main function"""
    try:
        if len(sys.argv) < 2:
            print("❌ Usage: python alert_contact_export.py <input_file.xlsx>")
            sys.exit(1)

        input_file = sys.argv[1]
        
        if not os.path.exists(input_file):
            print(f"❌ Input file not found: {input_file}")
            sys.exit(1)
        
        print(f"📁 Input file: {input_file}")
        
        # Clean the data
        cleaned_data = clean_data(input_file)
        
        # Generate output filename and path
        base_name = os.path.splitext(os.path.basename(input_file))[0]
        output_dir = r"C:\LocalAI\Stuart\broadly\Broadly Report\Alert Contact Export"
        
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        
        output_filename = os.path.join(output_dir, f"{base_name} - Alert Contact Export.xlsx")
        
        # Create the Excel file
        create_excel_file(cleaned_data, output_filename)
        
        print("✅ Data cleaning completed successfully!")
        print(f"📄 Output file: {output_filename}")

    except Exception as e:
        print(f"❌ Error occurred: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main() 